"""The transport service layer as ONE pipeline step (gen_v3 entry point).

Chains the three validated moves **in memory** on a loaded MARIO db —
no intermediate 900 MB exports (those belong to the standalone dev runs
of the individual scripts):

- **Move B** — splits "Other land transport" and "Transport via railways"
  into freight/passenger children, re-denominated to observed tkm/pkm;
- **Move A** — private mobility as household-operated activities
  (CAR.G/D/LPG/CNG/E + MOTO) producing "Private road mobility" (Mpkm),
  household motor fuels rerouted from Y, tailpipe re-attributed EY→E;
- **Move C** — own-account road freight externalised into its own
  activity + commodity, so the SUT perimeters match UNSD's (1221 = all
  road fuel in transport-family columns; industry rows keep process,
  heating and off-road).

Usage in gen_v3, after ``aggregate_ee`` + the steel/BFG block and before
the supply-mix updates (the new electricity input of CAR.E then joins
the electricity pooling for free)::

    from transport.pipeline import apply_transport_layer
    apply_transport_layer(db)

Standalone validation run (loads the base table, applies all three,
prints the acceptance numbers)::

    unset VIRTUAL_ENV; caffeinate -is \\
        /opt/anaconda3/envs/mario/bin/python transport/pipeline.py
"""

from __future__ import annotations

import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
sys.path.insert(0, str(HERE))

from apply_movea_write import TECH_ACT  # noqa: E402
from apply_movea_write import apply as _apply_a  # noqa: E402
from apply_moveb_split_write import CHILD_DEF  # noqa: E402
from apply_moveb_split_write import apply as _apply_b  # noqa: E402
from apply_movec_write import ACT as OWN_ACT  # noqa: E402
from apply_movec_write import apply as _apply_c  # noqa: E402


# The five split parents end up with zero output. Folding them away needs a
# target with the SAME unit — MARIO refuses to aggregate across units — and
# now that every transport child is physical (pkm or tkm), the only monetary
# transport commodity left is the residual category (63), so all five parents
# fold there. Numerically a no-op: they are zero.
EMPTY_PARENT_FOLD = {
    "Activity": {
        p: "Supporting and auxiliary transport activities; "
           "activities of travel agencies (63)"
        for p in ("Other land transport", "Transport via railways",
                  "Sea and coastal water transport", "Inland water transport",
                  "Air transport (62)")
    },
    "Commodity": {
        p: "Supporting and auxiliary transport services; travel agency services (63)"
        for p in ("Other land transportation services",
                  "Railway transportation services",
                  "Sea and coastal water transportation services",
                  "Inland water transportation services",
                  "Air transport services (62)")
    },
}


def fold_empty_parents(db) -> None:
    """Aggregate the zero-output split parents into same-unit targets."""
    import openpyxl  # noqa: PLC0415

    x = db.X.iloc[:, 0]
    lvl2 = db.X.index.get_level_values(2)
    left = {name: float(x[lvl2 == name].sum())
            for names in EMPTY_PARENT_FOLD.values() for name in names}
    hot = {n: v for n, v in left.items() if abs(v) > 1e-6}
    if hot:
        print(f"fold saltato: parent non vuoti {hot}", flush=True)
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in ("Activity", "Commodity", "Factor of production",
                  "Satellite account", "Consumption category", "Region"):
        ws = wb.create_sheet(sheet)
        ws.cell(row=1, column=2, value="Aggregation")
        for i, item in enumerate(db.get_index(sheet), start=2):
            ws.cell(row=i, column=1, value=item)
            target = EMPTY_PARENT_FOLD.get(sheet, {}).get(item)
            if target:
                ws.cell(row=i, column=2, value=target)
    path = HERE / "out" / "_fold_empty_parents.xlsx"
    path.parent.mkdir(exist_ok=True)
    wb.save(path)
    db.aggregate(str(path), ignore_nan=True)
    print(f"parent vuoti aggregati: {len(left)} item rimossi dalla griglia",
          flush=True)


def apply_transport_layer(db, validate: bool = False):
    """Apply Moves B, A and C in place; returns the db.

    ``validate=False`` (the pipeline default) skips the per-move
    ``calc_ghg`` checks: the footprint calculation belongs downstream,
    once every move has landed.
    """
    print("=== transport layer: Move B (split commercial transport) ===", flush=True)
    _apply_b(db)
    print("=== transport layer: Move A (private mobility) ===", flush=True)
    _apply_a(db, validate=validate)
    print("=== transport layer: Move C (own-account freight) ===", flush=True)
    _apply_c(db, write=True, validate=validate)
    print("=== transport layer: fold dei parent vuoti ===", flush=True)
    fold_empty_parents(db)
    print("=== transport layer: done ===", flush=True)
    return db


def main() -> None:
    """Standalone validation: base table -> B+A+C -> acceptance numbers."""
    import mario  # noqa: PLC0415
    import yaml  # noqa: PLC0415

    pfile = ROOT / ("paths_personal.yml" if (ROOT / "paths_personal.yml").exists()
                    else "paths.yml")
    paths = yaml.safe_load(open(pfile))["USER"]
    print("loading base table…", flush=True)
    db = mario.parse_from_txt(paths["raw"], table="SUT", mode="flows")
    db.aggregate(str(ROOT / "support" / "aggregate_ee.xlsx"), ignore_nan=True)

    apply_transport_layer(db)

    X = db.X.iloc[:, 0]
    lvl2 = db.X.index.get_level_values(2)

    def gx(name: str) -> float:
        return float(X[lvl2 == name].sum())

    def rx(region: str, name: str) -> float:
        return float(X[(db.X.index.get_level_values(0) == region) & (lvl2 == name)].sum())

    print("\n--- accettazione (attesi dai run standalone) ---", flush=True)
    print(f"IT ROAD.FRT   = {rx('IT', CHILD_DEF['ROAD.FRT'][0]):>12,.0f} Mtkm "
          f"(atteso ~127.800)", flush=True)
    priv = sum(gx(a) for a in TECH_ACT.values())
    print(f"mondo privato = {priv:>12,.0f} Mpkm (atteso ~16.030.000)", flush=True)
    print(f"IT privato    = {sum(rx('IT', a) for a in TECH_ACT.values()):>12,.0f} Mpkm "
          f"(atteso ~566.746)", flush=True)
    own = gx(OWN_ACT)
    hire = gx(CHILD_DEF["ROAD.FRT"][0])
    print(f"mondo own-acc = {own:>12,.0f} Mtkm (atteso ~2.165.734; quota "
          f"{100 * own / (own + hire):.1f}% vs 15,4% osservato)", flush=True)


if __name__ == "__main__":
    main()
