"""Move B (c), stage 2b — write the split into the table (add_sectors + surgery).

The write pass of the deterministic split validated by stage 2a:

1. **Register** the four children in the grid via MARIO's standard
   inventory path (`read_add_sectors_excel` + `add_sectors()`) with a
   register-only template generated from the blastfurnacegas.xlsx base
   (Master rows swapped, empty inventory sheets, DB-units extended) —
   NOT the `split=True`/cvxlab path (IOT-only machinery; decision
   2026-08-02: deterministic, MARIO untouched).
2. **Surgery** (the BFG idiom): per region, overwrite the children's
   coefficient columns (u/v/e = parent absolutes x shares ÷ child output)
   and the child-commodity rows (use rule + IPF, physical units), zero
   the parent, `update_scenarios('baseline', z, v, e, Y)` +
   `reset_to_coefficients`.
3. **Re-denomination**: child outputs are physical — X = Q (Mtkm/Mpkm)
   observed; where Q is unobserved it is synthesised as M ÷ median
   table-implied price of the child type (from the stage-2a dry-run CSV;
   the grid needs uniform units per commodity — declared per region).
   Rail/road children only; pipelines are already their own MEUR sector.
4. **Checks**: parent output gone, children alive, per-region commodity
   balance (supply == use) on the children, IT spot values; then
   `to_txt` export of the split baseline for inspection/reload.

Run:  unset VIRTUAL_ENV; caffeinate -is \
      /opt/anaconda3/envs/mario/bin/python transport/apply_moveb_split_write.py
"""

from __future__ import annotations

import csv
import shutil
import statistics
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
import yaml

HERE = Path(__file__).parent
ROOT = HERE.parent

# NXTR v0 recipe constants for the bottom-up fuel split; kept in step with
# build_moveb_split_spec, which declares them for the spec's provenance
INT_HGV, INT_BUS = 2.7e-4, 2.5e-4       # t fuel per vehicle-km
LOAD_DEFAULT, OCC_BUS = 10.0, 15.0      # tkm/vkm, pkm/vkm
PAX_TONNE = 0.1                         # one passenger = 100 kg of payload

PARENTS = {
    "road_pipe": ("Other land transport", "Other land transportation services"),
    "rail": ("Transport via railways", "Railway transportation services"),
    "sea": ("Sea and coastal water transport",
            "Sea and coastal water transportation services"),
    "iww": ("Inland water transport", "Inland water transportation services"),
    "air": ("Air transport (62)", "Air transport services (62)"),
}
# child key -> (activity name, commodity name, unit, inventory sheet)
CHILD_DEF = {
    "ROAD.FRT": ("Road freight transport", "Road freight transport services", "Mtkm", "RFT"),
    "ROAD.PAX": ("Road passenger transport", "Road passenger transport services", "Mpkm", "RPX"),
    "TRN.P": ("Rail passenger transport", "Rail passenger transport services", "Mpkm", "RLP"),
    "TRN.F": ("Rail freight transport", "Rail freight transport services", "Mtkm", "RLF"),
    "SEA.FRT": ("Sea and coastal freight transport",
                "Sea and coastal freight transport services", "Mtkm", "SEF"),
    "SEA.PAX": ("Sea and coastal passenger transport",
                "Sea and coastal passenger transport services", "Mpkm", "SEP"),
    "IWW.FRT": ("Inland water freight transport",
                "Inland water freight transport services", "Mtkm", "IWF"),
    "IWW.PAX": ("Inland water passenger transport",
                "Inland water passenger transport services", "Mpkm", "IWP"),
    "AIR.FRT": ("Air freight transport", "Air freight transport services",
                "Mtkm", "AIF"),
    "AIR.PAX": ("Air passenger transport", "Air passenger transport services",
                "Mpkm", "AIP"),
}
CHILDREN = {"road_pipe": ["ROAD.FRT", "ROAD.PAX"], "rail": ["TRN.P", "TRN.F"],
    "sea": ["SEA.FRT", "SEA.PAX"], "iww": ["IWW.FRT", "IWW.PAX"],
    "air": ["AIR.FRT", "AIR.PAX"],
}
# no open pkm exists for these: they keep the monetary denomination
# every transport child is now physical: air passenger-km derive
# from ICAO, sea passenger-km come from Eurostat, and inland
# waterway passengers — which no statistical system collects, by
# the explicit exclusion in Regulation (EC) 1365/2006 — are scaled
# with the sea passenger price (declared proxy, tiny sector).
MONETARY_CHILDREN: set[str] = set()
PRICE_PROXY = {"IWW.PAX": "SEA.PAX"}
FUEL_NAMES = [
    "Motor Gasoline", "Gas/Diesel Oil", "Liquefied Petroleum Gases (LPG)",
    "Biogasoline", "Biodiesels", "Other Liquid Biofuels", "Kerosene",
    "Heavy Fuel Oil", "Natural Gas Liquids",
]
TEMPLATE_BASE = ROOT / "support" / "add_sectors" / "blastfurnacegas.xlsx"
TEMPLATE_OUT = HERE / "out" / "_moveb_add_sectors.xlsx"
EXPORT_DIR = HERE / "out" / "_moveb_split_table"


def load_spec() -> dict:
    spec: dict = defaultdict(dict)
    with open(HERE / "data" / "moveb_split_spec.csv") as f:
        for r in csv.DictReader(f):
            if r["child"] == "PIPE":
                continue
            spec[(r["region"], r["block"], r["row_class"])][r["child"]] = float(r["share"])
    for (_, _, row_class), shares in spec.items():
        if row_class in ("other", "fuel_liquid"):
            tot = sum(shares.values())
            if tot > 0:
                for c in shares:
                    shares[c] /= tot
    return spec


def median_prices() -> dict[str, float]:
    """Median table-implied price per child (stage-2a dry-run) — the
    synthesiser for regions without observed Q."""
    vals: dict[str, list[float]] = defaultdict(list)
    with open(HERE / "data" / "moveb_split_dryrun.csv") as f:
        for r in csv.DictReader(f):
            if r["implied_price"]:
                vals[r["child"]].append(float(r["implied_price"]))
    return {c: statistics.median(v) for c, v in vals.items()}


def hgv_load_factors() -> dict[str, float]:
    """Per-country observed HGV load factor (tkm/vkm), for the road split."""
    load: dict[str, float] = {}
    with open(HERE / "data" / "derived_recipes_v0.csv") as f:
        for r in csv.DictReader(f):
            if (r["tech"] == "HGV" and r["coef"] == "load_factor_total"
                    and r["year"] == "Y11"):
                load[r["country"]] = float(r["value"])
    return load


def fuel_shares(block: str, children: list[str], Q: dict[str, float],
                region: str, load: dict[str, float]) -> dict[str, float]:
    """Split the parent's liquid fuel between the children by PHYSICAL WORK.

    Computed here rather than in the spec because here Q is final — observed
    where a statistic exists, synthesised from the child's revenue and its
    median implied price where none does. The spec could only see the
    observed volumes, and a volume that is merely *unmeasured* read there as
    zero activity: with one sibling missing the split degenerated to 0/1 and
    the whole of a country's marine bunker landed on its passenger child.
    Germany is the case in point — ITF publishes no coastal freight tonne-km
    for it, so its sea passenger sector came out at 128.021 gCO2eq/pkm on
    744 Mpkm. 33 of 48 regions were in that position for sea freight.

    Rules, unchanged in substance:
      - road: vehicle-km x NXTR fuel intensity (HGV against the country's
        observed load factor, bus against a default occupancy);
      - rail: traffic units, one intensity per unit at this stage;
      - water and air: tonne-km-equivalent, a passenger counting as 100 kg
        with baggage (ICAO DATA+, IATA RP 1726, EN 16258, GLEC). Splitting
        by revenue instead would hand ferries' and airlines' passenger side
        most of the fuel, since passenger revenue per unit of physical work
        is an order of magnitude above freight's.
    """
    if block == "road_pipe":
        w = {"ROAD.FRT": Q.get("ROAD.FRT", 0.0) / load.get(region, LOAD_DEFAULT) * INT_HGV,
             "ROAD.PAX": Q.get("ROAD.PAX", 0.0) / OCC_BUS * INT_BUS}
    elif block == "rail":
        w = {c: Q.get(c, 0.0) for c in children}
    else:
        w = {c: Q.get(c, 0.0) * (PAX_TONNE if c.endswith("PAX") else 1.0)
             for c in children}
    tot = sum(w.values())
    if tot <= 0:
        return {}
    return {c: v / tot for c, v in w.items()}


def ipf(seed: np.ndarray, rt: np.ndarray, ct: np.ndarray, iters: int = 30) -> np.ndarray:
    a = np.array(seed, dtype=float)
    for _ in range(iters):
        rs = a.sum(axis=1)
        nz = rs > 0
        a[nz] *= (rt[nz] / rs[nz])[:, None]
        cs = a.sum(axis=0)
        nz = cs > 0
        a[:, nz] *= ct[nz] / cs[nz]
    return a


def build_template(regions: list[str]) -> None:
    import openpyxl

    shutil.copy(TEMPLATE_BASE, TEMPLATE_OUT)
    wb = openpyxl.load_workbook(TEMPLATE_OUT)
    ws = wb["Master"]
    ws.delete_rows(2, ws.max_row - 1)
    for i, (act, com, unit, sheet) in enumerate(CHILD_DEF.values(), start=2):
        for j, val in enumerate(["GLOBAL", act, com, sheet, 1, unit, 1, None], start=1):
            ws.cell(row=i, column=j, value=val)
    for old in ("BFG", "O2G"):
        del wb[old]
    for _, (_, _, _, sheet) in CHILD_DEF.items():
        s = wb.create_sheet(sheet)
        for j, h in enumerate(["Quantity", "Unit", "Input", "Item type", "DB Item",
                               "DB Region", "Change type", "Source"], start=1):
            s.cell(row=1, column=j, value=h)
    units = wb["DB units"]
    row = units.max_row + 1
    for act, com, unit, _ in CHILD_DEF.values():
        units.cell(row=row, column=1, value="Activity")
        units.cell(row=row, column=2, value=act)
        units.cell(row=row, column=3, value="None")
        units.cell(row=row + 1, column=1, value="Commodity")
        units.cell(row=row + 1, column=2, value=com)
        units.cell(row=row + 1, column=3, value=unit)
        row += 2
    wb.save(TEMPLATE_OUT)
    print(f"template register-only -> {TEMPLATE_OUT}", flush=True)


def apply(db) -> None:
    """Move B on an already-loaded db, in place — the pipeline entry point."""
    regions = list(db.get_index("Region"))

    build_template(regions)
    db.read_add_sectors_excel(str(TEMPLATE_OUT), read_inventories=True)
    db.add_sectors()
    print("children registrati:", list(db.new_activities), flush=True)

    spec = load_spec()
    prices = median_prices()
    load = hgv_load_factors()
    print("prezzi mediani (sintesi Q mancanti):",
          {k: round(v, 3) for k, v in prices.items()}, flush=True)

    U, V, E, S, Y = db.U, db.V, db.E, db.S, db.Y
    X = db.X
    x_series = X.iloc[:, 0] if hasattr(X, "columns") else X  # first col, name-agnostic
    print(f"X shape: {getattr(X, 'shape', '?')}, col: "
          f"{list(getattr(X, 'columns', ['-']))[:2]}", flush=True)
    u, s, v, e = db.u, db.s, db.v, db.e
    coms = list(db.get_index("Commodity"))
    fuels = [c for c in FUEL_NAMES if c in coms]

    # Two-phase surgery: compute everything into dicts (fast read loop),
    # then write in bulk — column assignments on the natural axis, row
    # assignments via ONE transpose round-trip per matrix. Repeated .loc
    # row/scalar writes on the 70M-cell coefficient frames consolidate the
    # whole block per call under pandas-3 CoW (~hours); this path is ~two
    # big copies per matrix.
    synth = 0
    implausible: list[tuple] = []
    u_cols: dict[tuple, np.ndarray] = {}
    v_cols: dict[tuple, np.ndarray] = {}
    e_cols: dict[tuple, np.ndarray] = {}
    u_rows: dict[tuple, pd.Series] = {}
    y_rows: dict[tuple, np.ndarray] = {}
    s_rows: dict[tuple, pd.Series] = {}

    dropped_flow = 0.0        # use by consumers with non-positive output
    # Pass 1 — block arithmetic (monetary totals, physical outputs, shares)
    # for EVERY region first. A child's use row lands on parent activity
    # columns of other blocks AND other regions (Greek shipping is bought by
    # the German sea transport sector); those columns are about to be zeroed,
    # so each flow has to be re-pointed at the children that replace that
    # parent, which means knowing their Q up front. Without this the whole
    # transport-buys-transport block evaporates at rebuild.
    blocks: dict[tuple, dict] = {}
    for i, region in enumerate(regions):
        print(f"  [{i + 1}/{len(regions)}] {region} (calcolo)", flush=True)
        for block, (p_act, p_com) in PARENTS.items():
            children = CHILDREN[block]
            sh_other = spec[(region, block, "other")]
            sh_fuel = spec[(region, block, "fuel_liquid")]
            q_spec = spec.get((region, block, "Q"), {})

            u_col = U[(region, "Activity", p_act)]
            v_col = V[(region, "Activity", p_act)]
            e_col = E[(region, "Activity", p_act)]
            fuel_mask = u_col.index.get_level_values(2).isin(fuels)
            u_row = U.loc[(region, "Commodity", p_com), :]
            y_row = Y.loc[(region, "Commodity", p_com), :]
            users = pd.concat([u_row, y_row])
            M = float(users.sum())

            # transport self-use (subcontracting: the parent buys its own
            # commodity) would land on dead cells (zeroed parent row x
            # zeroed parent column) and vanish at rebuild — handled
            # explicitly as a child->child DIAGONAL block (freight
            # subcontracts freight): monetary self_val x share per child.
            self_col_key = (region, "Activity", p_act)
            self_row_key = (region, "Commodity", p_com)
            self_val = float(u_row[self_col_key])

            Q: dict[str, float] = {}
            for c in children:
                m_child = sh_other.get(c, 0.0) * M
                if c in MONETARY_CHILDREN:      # identity: output stays MEUR
                    Q[c] = m_child
                    continue
                q = q_spec.get(c, 0.0)
                if q > 0 and c in prices and m_child > 0:
                    # The observed volume ALWAYS wins: the physical output must
                    # match the statistics. A wildly off implied price is
                    # (almost always) the MONETARY side misbehaving — SBS
                    # coverage, transit, a sector that also does other things —
                    # so substituting Q with M / median price would replace a
                    # good observation with a value derived from the suspect
                    # one. Flag it for the radar instead.
                    price = m_child / q
                    if not (prices[c] / 4 <= price <= prices[c] * 4):
                        implausible.append((region, c, round(price, 4)))
                if q <= 0:
                    proxy = PRICE_PROXY.get(c)
                    if c not in prices and proxy in prices:
                        prices[c] = prices[proxy]
                    if m_child > 0 and c not in prices:
                        # no observed Q and no median price to synthesise from:
                        # the dry-run has not been re-run for this child. Refuse
                        # rather than scale a physical unit by a monetary value.
                        raise SystemExit(
                            f"{c}: nessun prezzo mediano (rilancia il dry-run "
                            f"apply_moveb_split.py prima del write)")
                    q = m_child / prices[c] if m_child > 0 else 0.0
                    synth += 1
                Q[c] = q
            # the fuel split follows the physical work the table will carry,
            # so it uses these Q — never the spec's observed-only volumes,
            # which read an unmeasured sibling as zero activity
            sh_fuel = fuel_shares(block, children, Q, region, load) or sh_fuel
            blocks[(region, block)] = dict(
                region=region, p_act=p_act, p_com=p_com, children=children, M=M, Q=Q,
                sh_other=sh_other, sh_fuel=sh_fuel, self_val=self_val,
                u_col=u_col, v_col=v_col, e_col=e_col, fuel_mask=fuel_mask,
                u_row=u_row, y_row=y_row, users=users,
                self_col_key=self_col_key, self_row_key=self_row_key)

    # parent activity column -> the children that replace it, with the
    # monetary share and the physical output to divide by
    replace: dict[tuple, list[tuple[tuple, float, float]]] = {
        (b["region"], "Activity", b["p_act"]): [
            ((b["region"], "Activity", CHILD_DEF[c][0]),
             b["sh_other"].get(c, 0.0), b["Q"][c])
            for c in b["children"] if b["Q"].get(c, 0.0) > 0
        ]
        for b in blocks.values()
    }
    # the re-pointing as flat arrays: flow at src_pos moves to dst_pos scaled
    # by w = monetary share / physical output of the receiving child
    col_index = U.columns
    src_pos, dst_pos, w_arr = [], [], []
    for pkey, targets in replace.items():
        for ckey, sh, q in targets:
            src_pos.append(col_index.get_loc(pkey))
            dst_pos.append(col_index.get_loc(ckey))
            w_arr.append(sh / q)
    src_pos = np.array(src_pos, dtype=int)
    dst_pos = np.array(dst_pos, dtype=int)
    w_arr = np.array(w_arr, dtype=float)
    is_parent = np.zeros(len(col_index), dtype=bool)
    is_parent[np.unique(src_pos)] = True
    xu_all = x_series.reindex(col_index).to_numpy(dtype=float)
    # a coefficient is flow / output: with output <= 0 there is no honest one,
    # and dividing anyway flips its sign — a negative entry in A voids the
    # non-negativity of the Leontief inverse, so the negative outputs spread
    # table-wide. Those cells are dropped and reported, never divided.
    ok_mask = (~is_parent) & (xu_all > 0)
    bad_mask = (~is_parent) & ~(xu_all > 0)
    inv_x = np.zeros(len(col_index))
    inv_x[ok_mask] = 1.0 / xu_all[ok_mask]

    # Pass 2 — surgery
    for j, ((region, block), B) in enumerate(blocks.items()):
        if j % len(PARENTS) == 0:
            print(f"  [{j // len(PARENTS) + 1}/{len(regions)}] {region} (chirurgia)",
                  flush=True)
        p_act, p_com = B["p_act"], B["p_com"]
        children, M, Q = B["children"], B["M"], B["Q"]
        sh_other, sh_fuel = B["sh_other"], B["sh_fuel"]
        u_col, v_col, e_col = B["u_col"], B["v_col"], B["e_col"]
        fuel_mask, u_row, y_row = B["fuel_mask"], B["u_row"], B["y_row"]
        users, self_val = B["users"], B["self_val"]
        self_col_key, self_row_key = B["self_col_key"], B["self_row_key"]

        pax = next(c for c in children if c.endswith("PAX") or c == "TRN.P")
        frt = next(c for c in children if c.endswith("FRT") or c == "TRN.F")
        users2 = users.drop(self_col_key)          # self allocated apart
        neg = users2.clip(upper=0.0).to_numpy()
        pos = users2.clip(lower=0.0).to_numpy()
        is_final = np.array(
            [(lvl[1] != "Activity" and any(k in str(lvl[2]) for k in
              ("households", "non-profit", "government")))
             for lvl in users2.index], dtype=bool)
        rule = np.zeros((len(children), len(users2)))
        rule[children.index(pax), is_final] = pos[is_final]
        rule[children.index(frt), ~is_final] = pos[~is_final]
        shares_vec = np.array([sh_other.get(c, 0.0) for c in children])
        seed = 0.9 * rule + 0.1 * np.outer(shares_vec, pos)
        closed = ipf(seed, shares_vec * pos.sum(), pos) + np.outer(shares_vec, neg)
        # monetary row total per child = sh x (M - self) + sh x self = sh x M

        n_u = len(u_row) - 1                      # users2 U-part length
        u2_index = users2.index[:n_u]
        for k, c in enumerate(children):
            act, com, _, _ = CHILD_DEF[c]
            m_child = sh_other.get(c, 0.0) * M
            xq = Q[c] if Q[c] > 0 else max(m_child, 1e-9)
            col = u_col * sh_other.get(c, 0.0)
            col[fuel_mask] = u_col[fuel_mask] * sh_fuel.get(c, 0.0)
            col[self_row_key] = 0.0               # self handled on the row side
            u_cols[(region, "Activity", act)] = (col / xq).to_numpy()
            v_cols[(region, "Activity", act)] = (v_col * sh_other.get(c, 0.0) / xq).to_numpy()
            e_cols[(region, "Activity", act)] = (e_col * sh_fuel.get(c, 0.0) / xq).to_numpy()
            one_hot = pd.Series(0.0, index=S.columns)
            one_hot[(region, "Commodity", com)] = 1.0
            s_rows[(region, "Activity", act)] = one_hot
            # physical scale: child's own monetary total, not M
            scale = (Q[c] / m_child) if m_child > 0 else 0.0
            phys = closed[k] * scale
            u_flows = pd.Series(0.0, index=u_row.index)
            u_flows[u2_index] = phys[:n_u]
            fl = u_flows.to_numpy(dtype=float)
            dropped_flow += float(np.abs(fl[bad_mask]).sum())
            cf = fl * inv_x
            np.add.at(cf, dst_pos, fl[src_pos] * w_arr)
            # self diagonal: physical self flow / X_child = self_val / M. The
            # own parent column was held out of the IPF, so it contributes
            # nothing above and there is no double count here.
            cf[col_index.get_loc((region, "Activity", act))] += (
                self_val / M if M > 0 else 0.0)
            u_rows[(region, "Commodity", com)] = pd.Series(cf, index=u_row.index)
            y_rows[(region, "Commodity", com)] = phys[n_u:]

        zero_col = np.zeros(len(u_col))
        u_cols[(region, "Activity", p_act)] = zero_col
        v_cols[(region, "Activity", p_act)] = np.zeros(len(v_col))
        e_cols[(region, "Activity", p_act)] = np.zeros(len(e_col))
        s_rows[(region, "Activity", p_act)] = pd.Series(0.0, index=S.columns)
        u_rows[(region, "Commodity", p_com)] = pd.Series(0.0, index=u_row.index)
        y_rows[(region, "Commodity", p_com)] = np.zeros(len(y_row))

    print(f"calcolo completato (Q sintetizzate: {synth}; prezzi impliciti "
          f"fuori banda, solo segnalati: {len(implausible)} "
          f"{implausible[:5]}; flusso scartato su consumatori con output<=0: "
          f"{dropped_flow:,.1f}); write bulk…", flush=True)

    # columns: natural axis, single-shot per matrix
    for key, arr in u_cols.items():
        u[key] = arr
    for key, arr in v_cols.items():
        v[key] = arr
    for key, arr in e_cols.items():
        e[key] = arr
    # rows: one transpose round-trip per matrix
    uT = u.T
    for key, ser in u_rows.items():
        uT[key] = ser.to_numpy()
    u = uT.T
    sT = s.T
    for key, ser in s_rows.items():
        sT[key] = ser.to_numpy()
    s = sT.T
    yT = Y.T
    for key, arr in y_rows.items():
        yT[key] = arr
    Y = yT.T
    print("surgery completata", flush=True)
    z = db.z
    z.update(s)
    z.update(u)
    db.update_scenarios("baseline", z=z, v=v, e=e, Y=Y)
    db.reset_to_coefficients("baseline")

    # --- checks ---
    X2 = db.X
    for block, (p_act, _) in PARENTS.items():
        xp = float(X2.loc[(slice(None), "Activity", p_act), :].to_numpy().sum())
        print(f"parent '{p_act}' output post-split: {xp:.3e} (atteso ~0)", flush=True)
    for c, (act, com, unit, _) in CHILD_DEF.items():
        xc = float(X2.loc[(slice(None), "Activity", act), :].to_numpy().sum())
        print(f"child {c}: X globale = {xc:,.0f} {unit}", flush=True)
    it_frt = float(X2.loc[("IT", "Activity", CHILD_DEF["ROAD.FRT"][0]), :].to_numpy().sum())
    print(f"IT ROAD.FRT output = {it_frt:,.1f} Mtkm (atteso ~ Q osservato)", flush=True)


def main() -> None:
    """Standalone dev run: load the base table, apply Move B, export."""
    pfile = ROOT / ("paths_personal.yml" if (ROOT / "paths_personal.yml").exists()
                    else "paths.yml")
    paths = yaml.safe_load(open(pfile))["USER"]
    import mario  # noqa: PLC0415

    print("loading base table…", flush=True)
    db = mario.parse_from_txt(paths["raw"], table="SUT", mode="flows")
    db.aggregate(str(ROOT / "support" / "aggregate_ee.xlsx"), ignore_nan=True)
    apply(db)
    EXPORT_DIR.mkdir(exist_ok=True)
    db.to_txt(path=str(EXPORT_DIR), scenario="baseline")
    print(f"export -> {EXPORT_DIR}", flush=True)


if __name__ == "__main__":
    main()
