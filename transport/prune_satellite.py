"""Reduce the satellite to the GHG accounts, at the aggregation step.

EXIOBASE hybrid ships 350 satellite accounts. Three of them are what the
pipeline reads — MARIO's ``exiobase_hybrid`` GHG profile weighs fossil CO2,
CH4 and N2O — and nothing in the pipeline touches a non-GHG account. The
nowcast does not need the rest either: its balance constraint is
conservation per commodity row, which lives in U/S/V/Y, and its endogenous
variables are output, fuel use, final demand, value added and imports.

So the table keeps a GHG vertical: the seven Kyoto-basket species plus
biogenic CO2, in both emission families, and drops the other 336 accounts by
marking them ``unused`` in the aggregation map — which is where MARIO's
``drop`` argument already looks.

What this gives up, and it is worth being plain: the resource, waste,
stock-addition and crop-residue blocks are the material Merciai's mass
balance is closed against, so the table can no longer demonstrate that an
activity conserves mass, and water, land and material footprints are off the
table for this vintage. All of it is RECOVERABLE — the cut happens at build
time, the base table is untouched, and un-marking a row brings it back on
the next build.

The size saving is not the reason: the satellite is 3,7% of the export
against U and S at 94%.

    uv run python transport/prune_satellite.py [--restore]
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
BOOK = HERE.parent / "support" / "aggregate_ee.xlsx"
SHEET = "Satellite account"
# the Kyoto basket as EXIOBASE names it, plus biogenic CO2 (GWP 0 by
# convention, but carbon that has to stay visible). Kept in BOTH emission
# families: 'Emis_unreg_w' carries 86,9 Mt of fossil CO2 — 0,29% of the
# total — that MARIO's hybrid profile does not currently weigh. Dropping it
# would settle that silently; keeping it leaves the choice open.
SPECIES = ["Carbon dioxide, fossil", "Carbon dioxide, biogenic",
           "CH4", "N2O", "HFCs", "PFCs", "SF6"]
FAMILIES = ["air - Emiss", "air - Emis_unreg_w"]
KEEP = {f"{s} ({f})" for s in SPECIES for f in FAMILIES}


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--restore", action="store_true",
                    help="Clear the unused marks, bringing every account back.")
    ap.add_argument("--book", type=Path, default=BOOK)
    args = ap.parse_args()

    wb = load_workbook(args.book)
    ws = wb[SHEET]
    kept, dropped, missing = [], 0, set(KEEP)
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        missing.discard(str(name))
        if args.restore:
            ws.cell(row=row, column=2).value = None
            continue
        if str(name) in KEEP:
            ws.cell(row=row, column=2).value = str(name)      # explicit self
            kept.append(str(name))
        else:
            ws.cell(row=row, column=2).value = "unused"
            dropped += 1
    if missing and not args.restore:
        raise SystemExit(f"accounts to keep not found in the map: {sorted(missing)}")
    wb.save(args.book)
    if args.restore:
        print(f"ripristinati tutti i conti satellite in {args.book}")
        return
    print(f"tenuti {len(kept)} conti GHG, marcati 'unused' {dropped}")
    for k in sorted(kept):
        print(f"    {k}")


if __name__ == "__main__":
    main()
