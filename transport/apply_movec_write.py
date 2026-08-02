"""Move C — own-account road freight extraction (dry-run first, then --write).

Externalises the sectors' internal logistics as ONE activity per region
("Own-account road freight transport") producing a SEPARATE commodity
("Own-account road transport", Mtkm) that each sector buys back — the SNA
ancillary-activity externalisation, data-driven. Aligns the SUT with the
UNSD perimeters (1221 = all road fuel in transport-family columns; industry
rows keep process/heating + off-road only) and makes transport-sector tkm
match the statistics (hire + own = ITF).

Mechanics (v0, declared):

- regional total = spec own tkm (ITF observed / EU-share fallback / from
  the B child at apply time for RoW); fuel = tkm ÷ own load factor × NXTR
  HGV intensity; **diesel-only** (own fleets ~all diesel);
- allocation weights = **column diesel × sector propensity** (the declared
  table seeded by observed NST own-shares, off-road-dampened for
  mining/agriculture/construction — their tractor/haul-truck diesel must
  stay put, as UNSD keeps it in the industry rows);
- per-cell cap α = 0.8 (a column is never stripped bare) + waterfall
  redistribution; an un-hostable residual is NOT forced: it stays
  embedded, reported in the gap register, and the own tkm scale down
  accordingly (graceful degradation — the LP's 1221 band spans the sum);
- direct CO2 moves with the fuel (IPCC diesel EF, capped by the cell's
  satellite availability); VA and other inputs stay in the sectors
  (fuel-only, the Move-A rule).

Default run is a **DRY-RUN**: prints the per-region realism table (top
extracted sectors, extraction shares, gaps) and writes
``data/movec_dryrun.csv`` — inspect before ``--write``.

Run:  unset VIRTUAL_ENV; caffeinate -is \
      /opt/anaconda3/envs/mario/bin/python transport/apply_movec_write.py [--write]
"""

from __future__ import annotations

import csv
import shutil
import sys
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd

HERE = Path(__file__).parent
ROOT = HERE.parent

TABLE = HERE / "out" / "_transport_table" / "flows"
TEMPLATE_BASE = ROOT / "support" / "add_sectors" / "blastfurnacegas.xlsx"
TEMPLATE_OUT = HERE / "out" / "_movec_add_sectors.xlsx"

ACT = "Own-account road freight transport"
COM = "Own-account road transport"
DIESEL = "Gas/Diesel Oil"
EF_DIES = 3.17          # tCO2 / t
# class-dependent cap: fleet-dominated cells (propensity >= 0.8) can yield
# up to 0.8 of their diesel; everything else at most 0.5.
ALPHA_FLEET, ALPHA_BASE, FLEET_PROP = 0.8, 0.5, 0.8
OWN_SHARE_EU = 0.154    # for the apply-side RoW fallback (from the B child)
CO2_NAME = "Carbon dioxide, fossil (air - Emiss)"
# columns never extracted from: transport family, energy supply, households
EXCLUDE_PAT = ["transport", "Transport", "Private car", "Private motorcycle",
               "Electricity", "Coal", "Gas", "Hydro", "Nuclear", "Solar", "Wind",
               "Bioenergy", "Other Fossil", "Other Renewables", "Steam and hot",
               "Petroleum Refinery", "households"]


def load_propensity() -> list[tuple[float, str]]:
    rules = []
    with open(HERE / "data" / "movec_propensity.csv") as f:
        for r in csv.DictReader(f):
            rules.append((float(r["weight"]), r["pattern"]))
    return rules


def propensity(name: str, rules: list[tuple[float, str]]) -> float:
    for pat in EXCLUDE_PAT:
        if pat in name and not name.startswith("Manufacture"):
            pass  # exclusion handled by rules below returning 0 on no match
    for w, pat in rules:
        if pat in name:
            return w
    return 0.0


def build_template() -> None:
    import openpyxl

    shutil.copy(TEMPLATE_BASE, TEMPLATE_OUT)
    wb = openpyxl.load_workbook(TEMPLATE_OUT)
    ws = wb["Master"]
    ws.delete_rows(2, ws.max_row - 1)
    for j, val in enumerate(["GLOBAL", ACT, COM, "OAT", 1, "Mtkm", 1, None], start=1):
        ws.cell(row=2, column=j, value=val)
    for old in ("BFG", "O2G"):
        del wb[old]
    s = wb.create_sheet("OAT")
    for j, h in enumerate(["Quantity", "Unit", "Input", "Item type", "DB Item",
                           "DB Region", "Change type", "Source"], start=1):
        s.cell(row=1, column=j, value=h)
    units = wb["DB units"]
    row = units.max_row + 1
    units.cell(row=row, column=1, value="Activity")
    units.cell(row=row, column=2, value=ACT)
    units.cell(row=row, column=3, value="None")
    units.cell(row=row + 1, column=1, value="Commodity")
    units.cell(row=row + 1, column=2, value=COM)
    units.cell(row=row + 1, column=3, value="Mtkm")
    wb.save(TEMPLATE_OUT)


def main() -> None:
    write = "--write" in sys.argv
    import mario  # noqa: PLC0415

    print(f"loading table ({'WRITE' if write else 'DRY-RUN'})…", flush=True)
    db = mario.parse_from_txt(str(TABLE), table="SUT", mode="flows")
    regions = list(db.get_index("Region"))
    rules = load_propensity()

    if write:
        build_template()
        db.read_add_sectors_excel(str(TEMPLATE_OUT), read_inventories=True)
        db.add_sectors()
        print("registrata:", list(db.new_activities), flush=True)

    spec = {r["region"]: r for r in csv.DictReader(open(HERE / "data" / "movec_spec.csv"))}
    U, E, Y = db.U, db.E, db.Y
    u, s, v, e = (db.u, db.s, db.v, db.e) if write else (None, None, None, None)
    X = db.X
    x_series = X.iloc[:, 0]
    row_index = U.index
    diesel_rows = row_index[row_index.get_level_values(2) == DIESEL]
    assert len(diesel_rows) == len(regions), "righe diesel != regioni"
    co2_rows = [i for i in E.index if str(i) == CO2_NAME]
    assert len(co2_rows) == 1
    CO2_ROW = co2_rows[0]

    dry_rows: list[dict] = []
    u_new_cols: dict[tuple, pd.Series] = {}
    e_new_cols: dict[tuple, pd.Series] = {}
    s_new_rows: dict[tuple, pd.Series] = {}
    diesel_row_updates: dict[tuple, np.ndarray] = {}   # U rows (per origin)
    e_cell_updates: dict[tuple, float] = {}            # (CO2, sector col) -> new coeff
    com_rows: dict[tuple, pd.Series] = {}              # new commodity use row
    gap_reg: list[tuple[str, float]] = []

    # activity columns eligible per region (exclude transport family etc.)
    act_cols_by_region: dict[str, list] = defaultdict(list)
    prop_by_col: dict[tuple, float] = {}
    for c in U.columns:
        if c[1] != "Activity":
            continue
        name = str(c[2])
        if any(p in name for p in EXCLUDE_PAT):
            continue
        w = propensity(name, rules)
        if w > 0:
            act_cols_by_region[c[0]].append(c)
            prop_by_col[c] = w

    U_diesel = U.loc[diesel_rows, :]                   # 48 x cols, absolute
    for region in regions:
        sp = spec.get(region)
        if sp is None:
            continue
        own_tkm = float(sp["own_tkm"])
        if own_tkm <= 0:                               # RoW: from the B child
            xk = (region, "Activity", "Road freight transport")
            hire = float(x_series.get(xk, 0.0))
            own_tkm = hire * OWN_SHARE_EU / (1 - OWN_SHARE_EU)
        demand = own_tkm * 1e6 / float(sp["load_own"]) * float(sp["int_hgv"])  # t diesel

        cols = act_cols_by_region.get(region, [])
        diesel_S = {c: float(U_diesel[c].clip(lower=0).sum()) for c in cols}
        w_S = {c: diesel_S[c] * prop_by_col[c] for c in cols}
        tot_w = sum(w_S.values())
        extracted = dict.fromkeys(cols, 0.0)
        if tot_w > 0 and demand > 0:
            residual = demand
            for _ in range(6):                          # waterfall
                head = {c: (ALPHA_FLEET if prop_by_col[c] >= FLEET_PROP
                            else ALPHA_BASE) * diesel_S[c] - extracted[c]
                        for c in cols}
                open_w = {c: w_S[c] for c in cols if head[c] > 1e-9}
                sw = sum(open_w.values())
                if sw <= 0 or residual <= 1e-9:
                    break
                for c, wgt in open_w.items():
                    take = min(residual * wgt / sw, head[c])
                    extracted[c] += take
                residual = demand - sum(extracted.values())
        got = sum(extracted.values())
        gap = demand - got
        own_eff = own_tkm * (got / demand) if demand > 0 else 0.0
        if demand > 0 and gap / demand > 0.02:
            gap_reg.append((region, gap / demand))

        # realism table rows
        top = sorted(extracted.items(), key=lambda kv: -kv[1])[:8]
        for c, val in top:
            if val <= 0:
                continue
            dry_rows.append(dict(
                region=region, sector=str(c[2])[:55], extracted_t=round(val, 0),
                cell_share=round(val / diesel_S[c], 3) if diesel_S[c] else "",
                weight=prop_by_col[c]))

        if not write:
            continue

        # --- surgery bookkeeping ---
        act_key = (region, "Activity", ACT)
        com_key = (region, "Commodity", COM)
        new_col = pd.Series(0.0, index=row_index)
        e_col = pd.Series(0.0, index=E.index)
        for c, val in extracted.items():
            if val <= 0:
                continue
            cell = U_diesel[c].clip(lower=0)
            take = cell / float(cell.sum()) * val       # pro-rata origins
            for orig_row, amt in take.items():
                key = orig_row
                if key not in diesel_row_updates:
                    diesel_row_updates[key] = U.loc[key, :].to_numpy(dtype=float).copy()
                pos = U.columns.get_loc(c)
                diesel_row_updates[key][pos] -= amt
                new_col[key] += amt
            # direct CO2 moves with the fuel (capped by the cell satellite)
            dco2 = val * EF_DIES
            xs = float(x_series.get(c, 0.0))
            eav = float(E.loc[CO2_ROW, c])
            dmove = min(dco2, max(eav, 0.0))
            if xs > 0:
                e_cell_updates[(CO2_ROW, c)] = (eav - dmove) / xs
            e_col[CO2_ROW] += dmove
        xq = own_eff if own_eff > 0 else 1e-9
        u_new_cols[act_key] = new_col / xq
        e_new_cols[act_key] = e_col / xq
        one_hot = pd.Series(0.0, index=db.S.columns)
        one_hot[com_key] = 1.0
        s_new_rows[(region, "Activity", ACT)] = one_hot
        buy = pd.Series(0.0, index=U.columns)
        for c, val in extracted.items():
            if val > 0:
                xs = float(x_series.get(c, 0.0))
                buy[c] = (own_eff * val / got) / xs if xs > 0 and got > 0 else 0.0
        com_rows[com_key] = buy

    # --- dry-run output ---
    with open(HERE / "data" / "movec_dryrun.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["region", "sector", "extracted_t",
                                          "cell_share", "weight"])
        w.writeheader()
        w.writerows(dry_rows)
    print(f"\nregioni con gap > 2%: {len(gap_reg)} "
          f"{[(r, round(g, 2)) for r, g in gap_reg[:8]]}", flush=True)
    for region in ("IT", "DE", "PL"):
        sel = [r for r in dry_rows if r["region"] == region][:6]
        print(f"  {region}:")
        for r in sel:
            print(f"    {r['sector']:57} {r['extracted_t']:>12,.0f} t "
                  f"(quota cella {r['cell_share']})", flush=True)
    if not write:
        print("\nDRY-RUN completo — ispeziona data/movec_dryrun.csv, poi --write")
        return

    # --- bulk write ---
    print("write bulk…", flush=True)
    for key, ser in u_new_cols.items():
        u[key] = ser.to_numpy()
        v[key] = 0.0
        e[key] = e_new_cols[key].to_numpy()
    uT = u.T
    for key, arr in diesel_row_updates.items():
        x_users = x_series.reindex(U.columns).to_numpy(dtype=float)
        with np.errstate(divide="ignore", invalid="ignore"):
            coeff = np.where(x_users > 0, arr / x_users, 0.0)
        uT[key] = coeff
    for key, ser in com_rows.items():
        uT[key] = ser.to_numpy()
    u = uT.T
    sT = s.T
    for key, ser in s_new_rows.items():
        sT[key] = ser.to_numpy()
    s = sT.T
    for (row, col), coeff in e_cell_updates.items():
        e.loc[row, col] = coeff
    print("surgery completata", flush=True)

    z = db.z
    z.update(s)
    z.update(u)
    db.update_scenarios("baseline", z=z, v=v, e=e, Y=Y)
    db.reset_to_coefficients("baseline")

    X2 = db.X
    mask = X2.index.get_level_values(2) == ACT
    print(f"own-account: X globale = {float(X2.iloc[:, 0][mask].sum()):,.0f} Mtkm", flush=True)
    try:
        db.calc_ghg(profile="exiobase_hybrid")
        f = db.f
        for reg in ("IT", "DE", "PL"):
            colk = (reg, "Activity", ACT)
            tot = float(f.loc["GHG AR6 GWP-100", colk])
            x = float(X2.loc[colk].iloc[0])
            dirc = float(db.E.loc[CO2_ROW, colk]) / x if x > 0 else 0.0
            print(f"  footprint own-account {reg}: {tot:.1f} g/tkm (dir {dirc:.1f})",
                  flush=True)
    except Exception as ex:
        print(f"WARN validazione footprint: {ex}", flush=True)

    db.to_txt(path=str(HERE / "out" / "_transport_table"), scenario="baseline")
    print("export aggiornato -> out/_transport_table", flush=True)


if __name__ == "__main__":
    main()
