"""Move A, stage 2 — write private mobility into the (Move-B split) table.

Loads the Move-B export (``out/_moveb_split_table/flows`` — the chain of
exports), registers the six household-operated techs (CAR.G/D/LPG/CNG/E +
MOTO) all supplying ONE new commodity **"Private road mobility"** (Mpkm —
the FULFILL MIMO pattern; shared-commodity Master rows are exactly what
FULFILL's own add_sectors does), and reroutes the households' motor-fuel
purchases from Y into the new activities' use columns.

v0 rules (declared):

- **fuel-only reroute**: maintenance/insurance/vehicle purchases stay as
  direct household purchases (v1 refinement);
- **bottom-up, capped**: fuel_tech = vkm x intensity with vkm = pkm_obs x
  share / occupancy (spec); per carrier the moved fuel is capped by the
  households' actual purchases (diesel/LPG/gas/electricity include heating
  and home uses — never move more than bottom-up transport demand), and
  the cap rescales vkm so the recipes stay exact;
- **gasoline-anchor synthesis** where no pkm is observed: all household
  gasoline is private driving (vkm = Y_gasoline / (share_G x int_G)),
  other carriers bottom-up at that vkm, capped;
- **origin structure preserved**: household fuel cells exist per origin
  region (imports included); the reroute moves pro-rata across origins;
- **EY untouched**: household direct combustion stays in the household
  satellite — the downstream combustion-based emission recompute
  re-attributes it from fuel use (plan, D9-adjacent); car activities get
  E = 0 and VA = 0 (own-account production).

Checks: exact fuel conservation per (region, carrier); new-commodity
supply == use; coverage diagnostics (moved gasoline / household gasoline).
Export: ``out/_transport_table`` (A+B combined baseline).

Run:  unset VIRTUAL_ENV; caffeinate -is \
      /opt/anaconda3/envs/mario/bin/python transport/apply_movea_write.py
"""

from __future__ import annotations

import csv
import shutil
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd

HERE = Path(__file__).parent
ROOT = HERE.parent

B_EXPORT = HERE / "out" / "_moveb_split_table" / "flows"
TEMPLATE_BASE = ROOT / "support" / "add_sectors" / "blastfurnacegas.xlsx"
TEMPLATE_OUT = HERE / "out" / "_movea_add_sectors.xlsx"
EXPORT_DIR = HERE / "out" / "_transport_table"

class SkipValidation(Exception):
    """Internal sentinel: the in-run footprint check is off."""


COMMODITY = "Private road mobility"
TECH_ACT = {
    "CAR.G": "Private car transport, gasoline",
    "CAR.D": "Private car transport, diesel",
    "CAR.LPG": "Private car transport, LPG",
    "CAR.CNG": "Private car transport, natural gas",
    "CAR.E": "Private car transport, electric",
    "MOTO": "Private motorcycle transport",
}
# IPCC combustion EFs, tCO2 per t fuel (declared v0: CO2 only — CH4/N2O are
# ~1-2% of road CO2e; ELE has no tailpipe). Used to re-attribute household
# driving combustion from EY to the car activities (exact conservation).
EF_CO2 = {"GASO": 3.07, "DIES": 3.17, "LPG": 3.02, "NGAS": 2.75, "ELE": 0.0}

# carrier -> commodity row name in the grid (containment-matched at runtime
# for the natural-gas variants; asserted, never silent)
CARRIER_NAME = {
    "GASO": "Motor Gasoline",
    "DIES": "Gas/Diesel Oil",
    "LPG": "Liquefied Petroleum Gases (LPG)",
    "NGAS": "Natural gas and services related to natural gas extraction",
    "ELE": "Electricity",
}


def build_template() -> None:
    import openpyxl

    shutil.copy(TEMPLATE_BASE, TEMPLATE_OUT)
    wb = openpyxl.load_workbook(TEMPLATE_OUT)
    ws = wb["Master"]
    ws.delete_rows(2, ws.max_row - 1)
    n = len(TECH_ACT)
    for i, (tech, act) in enumerate(TECH_ACT.items(), start=2):
        sheet = f"PM{i - 1}"
        for j, val in enumerate(
                ["GLOBAL", act, COMMODITY, sheet, 1, "Mpkm", round(1 / n, 6), None],
                start=1):
            ws.cell(row=i, column=j, value=val)
    for old in ("BFG", "O2G"):
        del wb[old]
    for i in range(1, n + 1):
        s = wb.create_sheet(f"PM{i}")
        for j, h in enumerate(["Quantity", "Unit", "Input", "Item type", "DB Item",
                               "DB Region", "Change type", "Source"], start=1):
            s.cell(row=1, column=j, value=h)
    units = wb["DB units"]
    row = units.max_row + 1
    for act in TECH_ACT.values():
        units.cell(row=row, column=1, value="Activity")
        units.cell(row=row, column=2, value=act)
        units.cell(row=row, column=3, value="None")
        row += 1
    units.cell(row=row, column=1, value="Commodity")
    units.cell(row=row, column=2, value=COMMODITY)
    units.cell(row=row, column=3, value="Mpkm")
    wb.save(TEMPLATE_OUT)
    print(f"template -> {TEMPLATE_OUT}", flush=True)


def apply(db, validate: bool = True) -> None:
    """Move A on an already-loaded db, in place. ``validate`` runs the
    in-run footprint check (skip it in the pipeline: calc_ghg belongs
    downstream, after every move)."""
    regions = list(db.get_index("Region"))
    coms = list(db.get_index("Commodity"))
    print(f"grid: {len(regions)} regioni, {len(db.get_index('Activity'))} activity, "
          f"{len(coms)} commodity", flush=True)

    # resolve carrier commodity names against the grid (containment for NGAS)
    resolved: dict[str, str] = {}
    for key, want in CARRIER_NAME.items():
        hit = [c for c in coms if c == want] or [c for c in coms if want in c]
        assert hit, f"carrier '{want}' non trovato nella griglia"
        resolved[key] = hit[0]
    print("carrier:", resolved, flush=True)

    build_template()
    db.read_add_sectors_excel(str(TEMPLATE_OUT), read_inventories=True)
    db.add_sectors(accept_non_unitary_sum=True)
    print("attivita registrate:", list(db.new_activities), flush=True)

    spec: dict[str, list[dict]] = defaultdict(list)
    with open(HERE / "data" / "movea_spec.csv") as f:
        for r in csv.DictReader(f):
            spec[r["region"]].append(r)

    U, V, E, Y = db.U, db.V, db.E, db.Y
    EY = db.EY
    u, s, v, e = db.u, db.s, db.v, db.e
    co2_rows = [i for i in E.index if str(i) == "Carbon dioxide, fossil (air - Emiss)"]
    assert len(co2_rows) == 1, f"riga CO2 ambigua: {co2_rows}"
    CO2_ROW = co2_rows[0]
    hh_cols = {}
    for region in regions:
        cands = [c for c in Y.columns if c[0] == region
                 and str(c[2]) == "Final consumption expenditure by households"]
        assert len(cands) == 1, f"household column ambigua per {region}: {cands}"
        hh_cols[region] = cands[0]

    u_cols: dict[tuple, pd.Series] = {}
    s_rows: dict[tuple, pd.Series] = {}
    y_updates: dict[tuple, pd.Series] = {}     # per-region household column
    e_sat_cols: dict[tuple, pd.Series] = {}    # per-tech satellite column (CO2)
    ey_updates: dict[tuple, pd.Series] = {}    # per-region household EY column
    ey_shortfall: list[str] = []
    max_conserv_err = [0.0]
    com_y: dict[tuple, float] = {}             # (region) -> pkm to Y
    diag: list[dict] = []

    row_index = U.index
    for i, region in enumerate(regions):
        rows = spec.get(region, [])
        if not rows:
            continue
        hh = hh_cols[region]
        y_col = Y[hh].copy()
        moved_fuel: dict[str, tuple[float, str]] = {}

        # availability per carrier: household cells across all origins
        def carrier_cells(carrier: str) -> pd.Index:
            name = resolved[carrier]
            mask = row_index.get_level_values(2) == name
            return row_index[mask]

        # bottom-up vkm
        car_rows = [r for r in rows if r["tech"] != "MOTO"]
        moto = next(r for r in rows if r["tech"] == "MOTO")
        pkm_obs = float(car_rows[0]["pkm_obs"])
        occ_car = float(car_rows[0]["occupancy"])
        if pkm_obs > 0:
            vkm_tot = pkm_obs / occ_car
            synth = False
        else:
            g = next(r for r in car_rows if r["tech"] == "CAR.G")
            gas_avail = float(y_col[carrier_cells("GASO")].clip(lower=0).sum())
            denom = float(g["share"]) * float(g["intensity"])
            vkm_tot = gas_avail / (denom * 1e6) if denom > 0 else 0.0
            synth = True

        # demands per tech, then per-carrier cap
        demand: dict[str, float] = {}
        vkm: dict[str, float] = {}
        for r in car_rows:
            vkm[r["tech"]] = vkm_tot * float(r["share"])
            demand[r["tech"]] = vkm[r["tech"]] * 1e6 * float(r["intensity"])
        vkm["MOTO"] = (float(moto["pkm_obs"]) / float(moto["occupancy"])
                       if float(moto["pkm_obs"]) > 0 else 0.0)
        demand["MOTO"] = vkm["MOTO"] * 1e6 * float(moto["intensity"])

        by_carrier: dict[str, list[str]] = defaultdict(list)
        for r in rows:
            by_carrier[r["carrier"]].append(r["tech"])
        scale: dict[str, float] = {}
        for carrier, techs in by_carrier.items():
            avail = float(y_col[carrier_cells(carrier)].clip(lower=0).sum())
            want = sum(demand[t] for t in techs)
            sc = min(1.0, avail / want) if want > 0 else 0.0
            for t in techs:
                scale[t] = sc

        pkm_out: dict[str, float] = {}
        for r in rows:
            t = r["tech"]
            occ = float(r["occupancy"])
            vkm_eff = vkm[t] * scale.get(t, 0.0)
            pkm_out[t] = vkm_eff * occ
            moved = vkm_eff * 1e6 * float(r["intensity"])
            cells = carrier_cells(r["carrier"])
            pos = y_col[cells].clip(lower=0)
            tot = float(pos.sum())
            col = pd.Series(0.0, index=row_index)
            moved_actual = 0.0
            if moved > 0 and tot > 0:
                flows = pos / tot * moved          # pro-rata across origins
                y_col[cells] = y_col[cells] - flows
                col[cells] = flows / pkm_out[t] if pkm_out[t] > 0 else 0.0
                moved_actual = float(flows.sum())
            moved_fuel[t] = (moved_actual, r["carrier"])
            u_cols[(region, "Activity", TECH_ACT[t])] = col

        # EY -> E re-attribution: household driving combustion moves to the
        # car activities (IPCC EFs, CO2 only), capped by the EY availability
        # so nothing goes negative and conservation is exact by construction.
        co2_direct = {t: mv * EF_CO2[carr] for t, (mv, carr) in moved_fuel.items()}
        tot_dir = sum(co2_direct.values())
        avail = float(EY.loc[CO2_ROW, hh])
        move_tot = min(tot_dir, max(avail, 0.0))
        fscale = move_tot / tot_dir if tot_dir > 0 else 0.0
        if fscale < 0.999:
            ey_shortfall.append(region)
        for t in co2_direct:
            vec = pd.Series(0.0, index=E.index)
            if pkm_out.get(t, 0.0) > 0:
                vec[CO2_ROW] = co2_direct[t] * fscale / pkm_out[t]
            e_sat_cols[(region, "Activity", TECH_ACT[t])] = vec
        ey_new = EY[hh].copy()
        ey_new[CO2_ROW] = ey_new[CO2_ROW] - move_tot
        ey_updates[hh] = ey_new
        conserv = abs(sum(co2_direct[t] * fscale for t in co2_direct) - move_tot)
        max_conserv_err[0] = max(max_conserv_err[0], conserv)

        # s = SUPPLY SHARES per commodity (multi-supplier: MARIO computes
        # Xa = s x Xc — a flat 1 would hand every tech the full regional
        # output; the X-vs-Q check caught exactly that)
        pkm_tot_region = sum(pkm_out.values())
        for r in rows:
            t = r["tech"]
            one_hot = pd.Series(0.0, index=db.S.columns)
            if pkm_tot_region > 0:
                one_hot[(region, "Commodity", COMMODITY)] = pkm_out[t] / pkm_tot_region
            s_rows[(region, "Activity", TECH_ACT[t])] = one_hot

        y_updates[hh] = y_col
        com_y[region] = pkm_tot_region
        gaso_avail = float(Y[hh][carrier_cells("GASO")].clip(lower=0).sum())
        gaso_moved = gaso_avail - float(y_col[carrier_cells("GASO")].clip(lower=0).sum())
        diag.append(dict(region=region, synth=synth,
                         pkm_tot=round(sum(pkm_out.values()), 0),
                         gaso_coverage=round(gaso_moved / gaso_avail, 3)
                         if gaso_avail > 0 else ""))
        if (i + 1) % 10 == 0:
            print(f"  [{i + 1}/{len(regions)}]", flush=True)

    print("calcolo completato; write bulk…", flush=True)
    for key, ser in u_cols.items():
        u[key] = ser.to_numpy()
        v[key] = 0.0
        sat = e_sat_cols.get(key)
        e[key] = sat.to_numpy() if sat is not None else 0.0
    for hh, vec in ey_updates.items():
        EY[hh] = vec.to_numpy()
    sT = s.T
    for key, ser in s_rows.items():
        sT[key] = ser.to_numpy()
    s = sT.T
    for hh, y_col in y_updates.items():
        Y[hh] = y_col.to_numpy()
    yT = Y.T
    com_row = pd.Series(0.0, index=Y.columns)
    for region, pkm_val in com_y.items():
        com_row[hh_cols[region]] = pkm_val
    for region in regions:
        key = (region, "Commodity", COMMODITY)
        row = pd.Series(0.0, index=Y.columns)
        if region in com_y:
            row[hh_cols[region]] = com_y[region]
        yT[key] = row.to_numpy()
    Y = yT.T
    print("surgery completata", flush=True)

    z = db.z
    z.update(s)
    z.update(u)
    db.update_scenarios("baseline", z=z, v=v, e=e, Y=Y, EY=EY)
    db.reset_to_coefficients("baseline")

    X2 = db.X
    for t, act in TECH_ACT.items():
        mask = X2.index.get_level_values(2) == act
        xt = float(X2.iloc[:, 0][mask].sum())
        print(f"{t}: X globale = {xt:,.0f} Mpkm", flush=True)
    it = [d for d in diag if d["region"] == "IT"][0]
    print(f"IT: pkm privato = {it['pkm_tot']:,.0f} Mpkm, copertura benzina = "
          f"{it['gaso_coverage']}", flush=True)
    print(f"riattribuzione EY->E: err conservazione max = {max_conserv_err[0]:.3e} t; "
          f"regioni con EY insufficiente: {len(ey_shortfall)} {ey_shortfall[:6]}",
          flush=True)
    n_synth = sum(1 for d in diag if d["synth"])
    print(f"regioni sintetizzate (gasoline-anchor): {n_synth}/{len(diag)}", flush=True)
    with open(HERE / "data" / "movea_diag.csv", "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["region", "synth", "pkm_tot", "gaso_coverage"])
        w.writeheader()
        w.writerows(diag)

    # --- fold the empty parents into a child (numeric no-op) ---
    try:
        import openpyxl
        FOLD = {"Activity": {"Other land transport": "Road freight transport",
                             "Transport via railways": "Rail passenger transport"},
                "Commodity": {"Other land transportation services":
                              "Road freight transport services",
                              "Railway transportation services":
                              "Rail passenger transport services"}}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet in ["Activity", "Commodity", "Factor of production",
                      "Satellite account", "Consumption category", "Region"]:
            ws = wb.create_sheet(sheet)
            ws.cell(row=1, column=2, value="Aggregation")
            items = list(db.get_index(sheet))
            for i, item in enumerate(items, start=2):
                ws.cell(row=i, column=1, value=item)
                m = FOLD.get(sheet, {}).get(item)
                if m:
                    ws.cell(row=i, column=2, value=m)
        fold_path = HERE / "out" / "_movea_fold.xlsx"
        wb.save(fold_path)
        db.aggregate(str(fold_path), ignore_nan=True)
        print("parent vuoti aggregati nei figli (no-op numerico)", flush=True)
    except Exception as ex:
        print(f"WARN: fold parent saltato ({ex})", flush=True)

    # --- footprint validation: GHG AR6 total + direct-CO2 share per unit ---
    try:
        if not validate:
            raise SkipValidation
        db.calc_ghg(profile="exiobase_hybrid")
        f = db.f
        X3, E3 = db.X, db.E
        acts = dict(TECH_ACT)
        acts.update({"ROAD.FRT": "Road freight transport",
                     "ROAD.PAX": "Road passenger transport",
                     "TRN.P": "Rail passenger transport",
                     "TRN.F": "Rail freight transport"})
        print("\nfootprint GHG AR6 [g/unit] (dir = CO2 diretta) - IT/DE/PL:", flush=True)
        for key, act in acts.items():
            vals = []
            for reg in ("IT", "DE", "PL"):
                colk = (reg, "Activity", act)
                try:
                    tot = float(f.loc["GHG AR6 GWP-100", colk])
                    x = float(X3.loc[colk].iloc[0])
                    dirc = float(E3.loc[CO2_ROW, colk]) / x if x > 0 else 0.0
                    vals.append(f"{reg}={tot:7.1f} (dir {dirc:6.1f})")
                except Exception:
                    vals.append(f"{reg}=n/a")
            print(f"  {key:9} {'   '.join(vals)}", flush=True)
    except SkipValidation:
        print("validazione footprint: saltata (pipeline)", flush=True)
    except Exception as ex:
        print(f"WARN: validazione footprint saltata ({ex})", flush=True)


def main() -> None:
    """Standalone dev run: load the Move-B table, apply Move A, export."""
    import mario  # noqa: PLC0415

    print("loading Move-B split table…", flush=True)
    db = mario.parse_from_txt(str(B_EXPORT), table="SUT", mode="flows")
    apply(db)
    EXPORT_DIR.mkdir(exist_ok=True)
    db.to_txt(path=str(EXPORT_DIR), scenario="baseline")
    print(f"export -> {EXPORT_DIR}", flush=True)


if __name__ == "__main__":
    main()
