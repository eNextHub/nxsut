"""Fill the toy economy into input_data/input_data.xlsx (long format).

Prior economy (consistent, balances close exactly):
  x0 = steel 100 t, power 50 TJ, services 200 MEUR
  COAL all imported (im_sh=1); ELE domestic; STEEL 20% imported.
Observations pull +10% on fuel use / power output, GDP 190 -> 195.
"""

import os

import openpyxl

os.chdir(os.path.dirname(os.path.abspath(__file__)))

A = ["steel", "power", "services"]  # activities
C = ["STEEL", "SERV", "COAL", "ELE"]  # commodities
F = ["COAL", "ELE"]  # fuels (energy carriers)
BK = ["b1211", "b121"]
G = ["industry", "services_g"]

# table -> {key tuple -> value}; keys follow the sheet's coordinate columns.
mat = lambda rows, cols, m: {(r, c): m[i][j] for i, r in enumerate(rows) for j, c in enumerate(cols)}

DATA = {
    # (commodities x activities) supply coefficients
    "s_prior": mat(C, A, [[1, 0, 0],      # STEEL by steel
                          [0, 0, 1],      # SERV by services
                          [0, 0, 0],      # COAL: none domestic
                          [0, 1, 0]]),    # ELE by power
    # frozen non-carrier technology (fuel rows unused by the slice -> 0)
    "u0_nonfuel": mat(C, A, [[0.05, 0.1, 0.02],
                             [0.10, 0.2, 0.10],
                             [0, 0, 0],
                             [0, 0, 0]]),
    # prior carrier coefficients (fuels x activities)
    "u0_fuel": mat(F, A, [[0.60, 0.50, 0.00],
                          [0.02, 0.00, 0.10]]),
    # relative L1 weights = 1/max(prior flow, 1)
    "W_fuel": mat(F, A, [[1 / 60, 1 / 25, 1.0],
                         [1 / 2, 1.0, 1 / 20]]),
    "W_y": {("STEEL",): 1 / 30, ("SERV",): 1 / 150, ("COAL",): 1 / 2, ("ELE",): 1 / 27},
    "W_v": {("steel",): 1 / 30, ("power",): 1 / 40, ("services",): 1 / 120},
    "v0": {("steel",): 0.3, ("power",): 0.8, ("services",): 0.6},
    "im_sh": {("STEEL",): 0.2, ("SERV",): 0.0, ("COAL",): 1.0, ("ELE",): 0.0},
    # observed carrier use per bucket (+10% vs prior); ELE/b121 NOT observed
    "F_obs": mat(F, BK, [[66.0, 93.5],
                         [2.2, 0.0]]),
    "M_fuel": mat(F, BK, [[1, 1],
                          [1, 0]]),
    "x_obs": {("steel",): 0, ("power",): 55.0, ("services",): 0},
    "M_x": {("steel",): 0, ("power",): 1, ("services",): 0},
    "Y_obs": {("STEEL",): 0, ("SERV",): 0, ("COAL",): 0, ("ELE",): 29.0},
    "M_y": {("STEEL",): 0, ("SERV",): 0, ("COAL",): 0, ("ELE",): 1},
    "Y_prior": {("STEEL",): 30.0, ("SERV",): 150.0, ("COAL",): 2.0, ("ELE",): 27.0},
    "EXP": {("STEEL",): 64.8, ("SERV",): 10.0, ("COAL",): 0.0, ("ELE",): 1.0},
    "VA_tgt": {("industry",): 72.0, ("services_g",): 123.0},
    "GDP": {(): 195.0},
    "B_ires": mat(A, BK, [[1, 1],
                          [0, 1],
                          [0, 0]]),
    "G_va": mat(A, G, [[1, 0],
                       [1, 0],
                       [0, 1]]),
    "tol": {("eps_f",): 0.02, ("eps_x",): 0.02, ("eps_y",): 0.02,
            ("eps_g",): 0.03, ("w_u",): 1.0, ("w_y",): 1.0, ("w_v",): 1.0},  # w_* rows unused now
}

INERT = {"regions_Name", "years_Name", "id"}  # key cols not part of DATA keys

wb = openpyxl.load_workbook("model/input_data/input_data.xlsx")
for sheet, table in DATA.items():
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    val_ix = headers.index("values")
    key_ixs = [i for i, h in enumerate(headers) if h not in INERT and h != "values"]
    filled = 0
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        key = tuple(row[i].value for i in key_ixs)
        if key in table:
            row[val_ix].value = float(table[key])
            filled += 1
        elif () in table:  # scalar tables (GDP): single row, empty key
            row[val_ix].value = float(table[()])
            filled += 1
        else:
            raise SystemExit(f"{sheet}: no value for key {key}")
    print(f"  {sheet}: {filled} values")
wb.save("model/input_data/input_data.xlsx")
print("inputs filled")
