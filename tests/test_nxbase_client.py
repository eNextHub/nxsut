"""Round-trip acceptance tests for the nxbase API client (WP0).

The client must rebuild, from the query API alone, inputs numerically
identical to the raw artifacts the pipeline used to read directly:

- trade matrices == the legacy shock-format workbooks in support/;
- the EMBER snapshot == MARIO's own reduction of the yearly full release.

Tests skip transparently when the nxbase API (or a raw file) is not
available — they are a local acceptance gate, not a CI unit suite.
"""

from __future__ import annotations

from pathlib import Path
from urllib.error import URLError
from urllib.request import urlopen

import pandas as pd
import pytest

import nxbase_client as nxc

REPO = Path(__file__).resolve().parents[1]

_EMBER_FUELS = (
    "Bioenergy",
    "Coal",
    "Gas",
    "Hydro",
    "Nuclear",
    "Other Fossil",
    "Other Renewables",
    "Solar",
    "Wind",
)


def _api_up() -> bool:
    try:
        urlopen(f"{nxc.DEFAULT_API}/health", timeout=3)
        return True
    except (URLError, OSError):
        return False


def _skip_unless_api() -> None:
    if not _api_up():
        pytest.skip("nxbase API not reachable (uv run nxbase api in the nxbase checkout)")


def test_split_item() -> None:
    assert nxc.split_item("a_coal-EXX-IT-Y24") == ["coal", "EXX", "IT", "Y24"]
    assert nxc.split_item("c_ELE-AT-Y23") == ["ELE", "AT", "Y23"]
    assert nxc.split_item("f_GHG.AR6.Y100") == ["GHG.AR6.Y100"]


@pytest.mark.parametrize("year", (2023, 2024, 2025))
def test_trade_matrix_round_trips_legacy_workbook(year: int) -> None:
    _skip_unless_api()
    legacy_path = REPO / "support" / f"trades_{year}.xlsx"
    if not legacy_path.exists():
        pytest.skip(f"{legacy_path} not present")

    legacy = pd.read_excel(legacy_path, sheet_name="z")
    legacy = legacy[
        legacy["row sector"].astype(str).eq("Electricity supply")
        & legacy["column sector"].astype(str).eq("Electricity need")
    ]
    expected = legacy.pivot_table(
        index="row region", columns="column region", values="value", aggfunc="sum"
    )
    expected.index = expected.index.astype(str)
    expected.columns = expected.columns.astype(str)
    expected = expected.sort_index().sort_index(axis=1)
    expected.index.name = None
    expected.columns.name = None

    matrix = nxc.get_trade_matrix(year=year).sort_index().sort_index(axis=1)

    assert list(matrix.index) == list(expected.index)
    assert list(matrix.columns) == list(expected.columns)
    pd.testing.assert_frame_equal(matrix, expected, check_dtype=False, atol=1e-12)
    assert (matrix.sum(axis=0) - 1.0).abs().max() < 1e-9


def _raw_ember_path() -> Path | None:
    import yaml

    paths_file = REPO / "paths.yml"
    if not paths_file.exists():
        return None
    users = yaml.safe_load(paths_file.read_text()) or {}
    for cfg in users.values():
        candidate = Path(cfg.get("ember", ""))
        if candidate.exists():
            return candidate
    return None


def test_ember_snapshot_round_trips_the_raw_release() -> None:
    _skip_unless_api()
    raw_csv = _raw_ember_path()
    if raw_csv is None:
        pytest.skip("EMBER yearly full release not present (paths.yml)")

    # MARIO's _reduce_ember_raw_release, reimplemented
    frame = pd.read_csv(raw_csv)
    expected = frame.loc[
        (frame["Area type"] == "Country or economy")
        & (frame["Category"] == "Electricity generation")
        & (frame["Subcategory"] == "Fuel")
        & (frame["Unit"] == "TWh")
        & (frame["Variable"].isin(_EMBER_FUELS)),
        ["ISO 3 code", "Year", "Variable", "Value"],
    ].rename(columns={"ISO 3 code": "ISO3"})
    expected = expected.loc[expected["ISO3"].astype(str).str.strip() != "", :]
    expected["Year"] = expected["Year"].astype(int)
    expected["Value"] = pd.to_numeric(expected["Value"], errors="coerce")
    expected = expected.dropna(subset=["Value"])

    snapshot = nxc.get_ember_snapshot()

    # nxbase stores no zeros and the client refills the fuel grid, so compare
    # the two frames as aligned grids (absent == 0 on both sides).
    key = ["ISO3", "Year", "Variable"]
    a = snapshot.set_index(key)["Value"]
    b = expected.set_index(key)["Value"]
    union = a.index.union(b.index)
    a = a.reindex(union, fill_value=0.0)
    b = b.reindex(union, fill_value=0.0)
    assert (a - b).abs().max() < 1e-9
    # and every nonzero raw value must be present in the snapshot
    nonzero = expected.loc[expected["Value"] != 0]
    assert nonzero.set_index(key).index.isin(snapshot.set_index(key).index).all()


# --- add_sectors recipe round-trip (steel/H2 inventory) -----------------------

# The pipeline's cluster: the base-DB label an nxbase-resolved concept maps back
# to (satellites/factors resolve to canonical flow names; commodities keep their
# label). This lives here, not in nxbase, by design.
_ADD_SECTORS_CLUSTER = {
    "Carbon dioxide, fossil": "CO2",
    "Methane, fossil": "CH4",
    "Nitrous oxide": "N2O",
    "Operating surplus: Consumption of fixed capital": "CAPEX",
    "Coke Oven Coke": "Coke",
}
_STEEL_MASTER = REPO / "support" / "add_sectors" / "Master_steel_h2.xlsx"


def _master_recipe() -> dict[tuple[str, str], float]:
    """(activity, input label) -> summed quantity, from the original master."""
    import openpyxl

    wb = openpyxl.load_workbook(_STEEL_MASTER, read_only=True)
    ms = wb["Master"]
    mh = [c.value for c in next(ms.iter_rows(min_row=1, max_row=1))]
    mi = {h: i for i, h in enumerate(mh)}
    sheet_to_act = {
        r[mi["Inventory sheet"]]: r[mi["Activity"]]
        for r in ms.iter_rows(min_row=2, values_only=True)
        if r[mi["Inventory sheet"]] is not None
    }
    out: dict[tuple[str, str], float] = {}
    skip = {"Master", "Commodities Clusters", "Regions Clusters", "DB units"}
    for sheet in wb.sheetnames:
        if sheet in skip:
            continue
        ws = wb[sheet]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(hdr)}
        act = str(sheet_to_act.get(sheet, "")).strip()
        for row in ws.iter_rows(min_row=2, values_only=True):
            label = row[idx["Input"]]
            try:
                qty = float(row[idx["Quantity"]] or 0)
            except (TypeError, ValueError):
                qty = 0.0
            if label and qty:
                out[(act, str(label).strip())] = out.get((act, str(label).strip()), 0.0) + qty
    return out


def test_add_sectors_recipe_round_trips_the_master() -> None:
    _skip_unless_api()
    if not _STEEL_MASTER.exists():
        pytest.skip(f"{_STEEL_MASTER} not present")

    recipe = nxc.get_add_sectors_recipe()
    recon: dict[tuple[str, str], float] = {}
    for _, r in recipe[recipe["item_type"] != "output"].iterrows():
        label = _ADD_SECTORS_CLUSTER.get(r["input"], r["input"])  # resolved -> base label
        recon[(r["route"], label)] = recon.get((r["route"], label), 0.0) + r["quantity"]

    master = _master_recipe()
    # Every non-trivial recipe cell must match the master's quantity (tolerance
    # set by the CSV channel's ~6 significant figures, not a real deviation).
    for key, qty in master.items():
        if abs(qty) < 1e-7:  # below every real coefficient (~1e-6+): numerical noise
            continue
        got = recon.get(key, 0.0)
        assert abs(got - qty) <= 1e-4 * abs(qty) + 1e-9, f"{key}: nxbase {got} != master {qty}"

    # And every route produces its commodity (one SUP per route).
    assert (recipe["item_type"] == "output").sum() == 24
