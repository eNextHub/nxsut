"""Query-API client for nxbase — the pipeline's only data doorway.

nxbase exposes its PostgreSQL content through a query API (`/data`,
`/data.csv`, `/sets/*`); this module wraps the calls the gen_v*.ipynb
notebooks need and reshapes the rows into the exact structures MARIO
consumes. The translation
logic lives here (the consumer), the base is always the API — see
nxbase docs/knowledge/nxsut_bridge.md, decision 2026-07-13.

Data access: the hosted nxbase API serves the open sources anonymously (no
login, no key) — set `api_url` / `paths.yml:nxbase_api` to the public instance
(`https://enextgen.it/nxbase-api`); the nxbase repository itself is private.
Run a local nxbase (`uv run nxbase api`, Docker Postgres up) only if you have
the checkout, then point `api_url` at `http://127.0.0.1:8000`.
"""

from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import urlencode
from urllib.request import urlopen

import pandas as pd

DEFAULT_API = "http://127.0.0.1:8000"

# EMBER covers 2000+; nxbase periods are Y<yy> shorts.
_CENTURY = 2000
# User-assigned ISO codes country_converter may not resolve.
_ISO2_TO_ISO3_OVERRIDES = {"XK": "XKX"}  # Kosovo


def _get_json(api_url: str, path: str) -> object:
    with urlopen(f"{api_url.rstrip('/')}{path}") as resp:  # noqa: S310 (trusted host)
        return json.load(resp)


def _get_csv(api_url: str, params: dict[str, object]) -> pd.DataFrame:
    query = urlencode({k: v for k, v in params.items() if v is not None}, doseq=True)
    return pd.read_csv(f"{api_url.rstrip('/')}/data.csv?{query}")


def split_item(item: str) -> list[str]:
    """`a_coal-EXX-IT-Y24` -> ['coal', 'EXX', 'IT', 'Y24'].

    Shorts contain neither `_` nor `-` by nxbase convention, so the first
    `_` separates the type letter and `-` separates the attributes.
    """
    return item.split("_", 1)[1].split("-")


def _iso2_to_iso3(codes: pd.Series) -> pd.Series:
    import country_converter as coco

    unique = sorted(set(codes) - set(_ISO2_TO_ISO3_OVERRIDES))
    converted = coco.convert(names=unique, src="ISO2", to="ISO3", not_found=None)
    if isinstance(converted, str):  # coco returns a scalar for a single name
        converted = [converted]
    mapping = dict(zip(unique, converted)) | _ISO2_TO_ISO3_OVERRIDES
    return codes.map(mapping)


def get_provenance(api_url: str = DEFAULT_API, source_names: list[str] | None = None) -> str:
    """One-line-per-fact provenance stamp to print in the notebook run."""
    health = _get_json(api_url, "/health")
    lines = [f"nxbase version: {health['version']} (api: {api_url})"]
    if source_names:
        rows = _get_json(api_url, "/sets/source")
        for row in rows:
            if row.get("name") in source_names:
                lines.append(
                    f"source: {row.get('extended')} "
                    f"(version {row.get('version')}, status {row.get('status')})"
                )
    return "\n".join(lines)


def get_ember_snapshot(
    api_url: str = DEFAULT_API,
    source: str = "EMBER Yearly Electricity Data (generation) 2025",
) -> pd.DataFrame:
    """Electricity generation snapshot in MARIO's reduced EMBER format.

    Queries `/data.csv?parameter=Total output&source=...` and returns the
    4-column frame `update_supply_mix(ember_path=...)` reads natively
    (write it to a transient CSV): ISO3, Year, Variable, Value (TWh).
    """
    frame = _get_csv(api_url, {"parameter": "Total output", "source": source, "sort": "id"})
    if frame.empty:
        raise ValueError(f"no Total output rows for source {source!r}")

    parts = frame["item_1"].map(split_item)
    tech_short = parts.str[0]
    iso2 = parts.str[2]
    year = parts.str[3].str[1:].astype(int) + _CENTURY

    # exact EMBER labels live in the EMBER-classified activity rows
    activities = _get_json(api_url, "/sets/activity")
    label_by_short = {
        r["short"]: r["name"] for r in activities if r.get("classification") == "EMBER"
    }

    out = pd.DataFrame(
        {
            "ISO3": _iso2_to_iso3(iso2),
            "Year": year,
            "Variable": tech_short.map(label_by_short),
            "Value": frame["value"].astype(float),
        }
    )
    dropped = out["ISO3"].isna() | out["Variable"].isna()
    if dropped.any():
        print(f"nxbase_client: dropped {int(dropped.sum())} unmappable rows")
    out = out.loc[~dropped]

    # nxbase stores no zero rows (skip-zeros policy): rebuild the full
    # (country, year) x fuel grid so MARIO sees explicit zero shares.
    grid = (
        out.pivot_table(index=["ISO3", "Year"], columns="Variable", values="Value")
        .reindex(columns=sorted(label_by_short.values()))
        .fillna(0.0)
        .stack()
        .rename("Value")
        .reset_index()
    )
    return grid


_UNSD_GEN_SOURCE = "UNSD Energy Statistics — electricity & heat production 2021-2023"

# UNSD -> EMBER-family mapping (same declared mapping as the step-0 selector,
# nowcast/step0_electricity_source_selector.py). Non-thermal families from the
# plant-type source totals (015*/016* on SIEC 7000, main + autoproducer);
# thermal families from the by-fuel view (01<fuel> on SIEC 7000T). Pumped
# hydro (PH, an "of which" memo) is subtracted from Hydro — EMBER excludes
# pumped-storage output. Renewables/nuclear conventions never enter (observed
# generation, not primary equivalents).
_UNSD_PLANT_FAMILY = {
    "N": "Nuclear", "HY": "Hydro", "S": "Solar", "W": "Wind",
    "G": "Other Renewables", "T": "Other Renewables",
    "O": "Other Fossil", "H": "Other Fossil",
}
_UNSD_PUMPED = "PH"
_UNSD_FUEL_FAMILY = {
    "CL": "Coal", "CP": "Coal", "LB": "Coal",
    "NG": "Gas",
    "CR": "Other Fossil", "RF": "Other Fossil", "DL": "Other Fossil",
    "PP": "Other Fossil", "OS": "Other Fossil", "PT": "Other Fossil",
    "MG": "Other Fossil", "NRW": "Other Fossil",
    "BI": "Bioenergy", "SBF": "Bioenergy", "LBF": "Bioenergy",
    "BS": "Bioenergy", "RW": "Bioenergy",
}
_EMBER_LABELS = ["Bioenergy", "Coal", "Gas", "Hydro", "Nuclear",
                 "Other Fossil", "Other Renewables", "Solar", "Wind"]


def get_unsd_generation_snapshot(
    api_url: str = DEFAULT_API,
    source: str = _UNSD_GEN_SOURCE,
) -> pd.DataFrame:
    """UNSD electricity generation in MARIO's reduced EMBER format.

    Queries `/data.csv?parameter=Supply&source=UNSD.GEN...` and maps the
    plant-type + by-fuel views onto the 9 EMBER family labels, so the frame
    is drop-in for `update_supply_mix` (ISO3, Year, Variable, Value TWh).
    """
    frame = _get_csv(api_url, {"parameter": "Supply", "source": source, "sort": "id"})
    if frame.empty:
        raise ValueError(f"no Supply rows for source {source!r}")

    com = frame["item_1"].str[2:]                       # c_<SIEC>
    parts = frame["item_2"].map(split_item)             # [tx, iso2, Yxx]
    tx, iso2 = parts.str[0], parts.str[1]
    year = parts.str[2].str[1:].astype(int) + _CENTURY
    twh = frame["value"].astype(float) / 1e3            # GWh -> TWh

    producer = tx.str[:3]
    suffix = tx.str[3:]
    fam = pd.Series(pd.NA, index=frame.index, dtype="object")
    sign = pd.Series(1.0, index=frame.index)

    plant = (com == "7000") & producer.isin(("015", "016"))
    fam[plant] = suffix[plant].map(_UNSD_PLANT_FAMILY)
    pumped = plant & (suffix == _UNSD_PUMPED)
    fam[pumped], sign[pumped] = "Hydro", -1.0
    byfuel = com == "7000T"
    fam[byfuel] = tx[byfuel].str[2:].map(_UNSD_FUEL_FAMILY)

    keep = fam.notna()
    out = (
        pd.DataFrame({"ISO3": _iso2_to_iso3(iso2[keep]), "Year": year[keep],
                      "Variable": fam[keep], "Value": (twh * sign)[keep]})
        .dropna(subset=["ISO3"])
        .groupby(["ISO3", "Year", "Variable"], as_index=False)["Value"].sum()
    )
    negative = out["Value"] < 0
    if negative.any():                                   # PH memo > HY total: misreport
        print(f"nxbase_client: clipped {int(negative.sum())} negative family rows")
        out.loc[negative, "Value"] = 0.0

    grid = (
        out.pivot_table(index=["ISO3", "Year"], columns="Variable", values="Value")
        .reindex(columns=_EMBER_LABELS)
        .fillna(0.0)
        .stack()
        .rename("Value")
        .reset_index()
    )
    return grid


def get_supply_mix_snapshot(
    api_url: str = DEFAULT_API,
    years: tuple[int, ...] = (2021, 2022, 2023),
    selection_path: str | Path | None = None,
) -> pd.DataFrame:
    """Blended generation snapshot: UNSD-first, EMBER fallback (step 0).

    Per country, uses the UNSD.GEN family frame when the arbitrated step-0
    selection says so (``nowcast/step0_efficiency.csv``: selector TVD +
    implied-efficiency arbitration) and the country reports the requested
    year; every other (country, year) keeps the EMBER rows. Same 4-column
    shape as :func:`get_ember_snapshot` — drop-in for `update_supply_mix`.
    """
    selection_path = Path(selection_path) if selection_path else (
        Path(__file__).resolve().parent.parent / "nowcast" / "data" / "step0_efficiency.csv"
    )
    sel = pd.read_csv(selection_path)
    latest = sel.sort_values("year").groupby("site").tail(1)
    unsd_iso2 = latest.loc[
        latest["arbitrated_decision"].str.startswith("UNSD", na=False), "site"
    ]
    unsd_iso3 = set(_iso2_to_iso3(unsd_iso2).dropna())

    ember = get_ember_snapshot(api_url)
    ember = ember[ember["Year"].isin(years)]
    unsd = get_unsd_generation_snapshot(api_url)
    unsd = unsd[unsd["Year"].isin(years)]

    # empty-safe masks: a year outside UNSD coverage (2024+) must degrade to
    # pure EMBER for every country, selected or not.
    take_unsd = unsd["ISO3"].isin(unsd_iso3)
    taken = set(map(tuple, unsd.loc[take_unsd, ["ISO3", "Year"]].values))
    take_ember = pd.Series(
        [(i, y) not in taken for i, y in zip(ember["ISO3"], ember["Year"])],
        index=ember.index,
    )

    blended = pd.concat(
        [unsd.loc[take_unsd], ember.loc[take_ember]], ignore_index=True
    ).sort_values(["ISO3", "Year", "Variable"], ignore_index=True)

    n_unsd = len(taken)
    n_ember = ember.loc[take_ember, ["ISO3", "Year"]].drop_duplicates().shape[0]
    print(f"nxbase_client: blended supply mix — {n_unsd} (country, year) from "
          f"UNSD.GEN, {n_ember} from EMBER (selection: {selection_path.name})")
    return blended


def get_trade_matrix(
    api_url: str = DEFAULT_API,
    year: int = 2024,
    commodity: str = "Electricity",
    source: str | None = None,
) -> pd.DataFrame:
    """Origins x destinations share matrix for one commodity and year.

    Queries `/data.csv?parameter=Import mix&period=<year>&commodity=...` and
    pivots to the matrix `update_trade_mix` consumes (region shorts on both
    axes, every destination column summing to 1, domestic diagonal included).

    `source` pins the origin dataset by name (several import-mix sources now
    coexist in nxbase): the proprietary Electricity Maps set
    (`"Electricity Maps Import mix <year>"`, used by v2.x) or the open ENTSO-E
    physical set (`"ENTSO-E electricity import mix <year>"`, used by v3.0).
    Leave `None` only when a single source is present.
    """
    params: dict[str, object] = {
        "parameter": "Import mix",
        "period": str(year),
        "commodity": commodity,
        "sort": "id",
    }
    if source is not None:
        params["source"] = source
    frame = _get_csv(api_url, params)
    if frame.empty:
        raise ValueError(
            f"no Import mix rows for {commodity!r} in {year}"
            + (f" from source {source!r}" if source else "")
        )

    destination = frame["item_1"].map(lambda s: split_item(s)[1])
    origin = frame["item_2"].map(lambda s: split_item(s)[0])
    matrix = pd.DataFrame(
        {"origin": origin, "destination": destination, "value": frame["value"].astype(float)}
    ).pivot(index="origin", columns="destination", values="value")
    # nxbase stores no zero rows (skip-zeros policy): re-materialize them so
    # every destination column carries the full origin universe (a missing
    # origin must force a zero share in update_trade_mix, not subset-rescale).
    universe = sorted(set(matrix.index) | set(matrix.columns))
    matrix = matrix.reindex(index=universe, columns=universe).fillna(0.0)
    matrix.index.name = None
    matrix.columns.name = None
    return matrix


def get_add_sectors_recipe(
    api_url: str = DEFAULT_API,
    source: str = "Ghezzi et al. 2026 - steel & H2 inventory",
) -> pd.DataFrame:
    """Fetch an add_sectors unit-process recipe from nxbase, reshaped by route.

    Queries `/data.csv?source=<name>` and returns the per-route inventory the
    pipeline reattaches to a base table (via its own clusters/placement) before
    MARIO `add_sectors`. Columns: `route` (the new activity), `input` (the
    resolved item name), `item_type` (`Commodity` | `Satellite account` |
    `Factor of production` | `output`), `quantity`, `unit`.

    nxbase holds only the *generic recipe* (quantities + backbone-anchored
    items). The *base-DB attachment* — which EXIOBASE commodity each input maps
    to (the cluster / DB Item), the GLOBAL/market-share placement, the
    furnace-gas emission reallocation — stays here in the pipeline: the recipe's
    resolved concepts (e.g. `Carbon dioxide, fossil`) are mapped back to the
    base table's labels by the caller. Round-trip verified: the reconstructed
    quantities equal the original master.
    """
    frame = _get_csv(api_url, {"source": source, "sort": "id", "limit": 100000})
    if frame.empty:
        raise ValueError(f"no add_sectors rows for source {source!r}")

    def _item_type(row: pd.Series) -> str:
        if row["parameter"] == "VA":
            return "Factor of production"
        if row["parameter"] == "SUP":
            return "output"
        if row["i1_set"] == "flow":
            return "Satellite account"
        return "Commodity"

    out = pd.DataFrame(
        {
            "route": frame["i2_name"].fillna(frame["i1_name"]).astype(str).str.strip(),
            "input": frame["i1_name"].astype(str).str.strip(),
            "item_type": frame.apply(_item_type, axis=1),
            "quantity": frame["value"].astype(float),
            "unit": frame["unit"],
        }
    )
    return out.reset_index(drop=True)


_STEEL_ROUTE_LABEL = {"BF.BOF": "BF-BOF", "DRI.EAF": "DRI-EAF", "SCRAP.EAF": "scrap-EAF"}


def get_steel_route_mix(
    api_url: str = DEFAULT_API,
    year: int = 2024,
    source: str = "World Steel in Figures - crude steel by route",
) -> dict[str, dict[str, float]]:
    """Per-region crude-steel production shares by route, for `update_supply_mix`.

    Queries `/data.csv?parameter=Supply&source=WSTEEL&period=<year>` (SUP holds
    absolute Mt by route — nxbase ingests the level, the mix is derived here) and
    returns `{ISO2 region: {route: share}}`, shares summing to 1 per region. Routes:
    `BF-BOF` (primary, blast furnace), `DRI-EAF` (primary, ore via DRI), `scrap-EAF`
    (secondary/recycled).

    For the EXIOBASE 2-way split, collapse in the notebook: primary = BF-BOF + DRI-EAF,
    secondary = scrap-EAF, mapped to the base table's steel activities (primary steel
    vs the secondary re-processing activity). Region keys are ISO2 — map to the table's
    regions (44 EXIOBASE + RoW) at the call site.
    """
    frame = _get_csv(
        api_url, {"parameter": "Supply", "source": source, "period": str(year), "sort": "id"}
    )
    if frame.empty:
        raise ValueError(f"no Supply rows for source {source!r} in {year}")
    parts = frame["item_2"].map(split_item)  # [route_short, ISO2, period]
    df = pd.DataFrame(
        {
            "region": parts.str[1],
            "route": parts.str[0].map(_STEEL_ROUTE_LABEL),
            "value": frame["value"].astype(float),
        }
    )
    mix: dict[str, dict[str, float]] = {}
    for region, g in df.groupby("region"):
        total = g["value"].sum()
        if total > 0:
            mix[region] = {r: round(v / total, 6) for r, v in zip(g["route"], g["value"])}
    return mix


# base-table (EXIOBASE) activities competing to supply the steel commodity
STEEL_PRIMARY_ACT = "Manufacture of basic iron and steel and of ferro-alloys and first products thereof"
STEEL_SECONDARY_ACT = "Re-processing of secondary steel into new steel"
STEEL_COMMODITY = "Basic iron and steel and of ferro-alloys and first products thereof"


def _iso2_to_exiobase_region(iso2s, regions) -> dict[str, str | None]:
    """ISO2 -> table region: the 44 EXIOBASE members kept, others -> RoW (WA/WE/WF/WL/WM)."""
    from mario.clusters.coverage import load_exiobase_region_members

    keep = set(regions) if regions is not None else set()
    iso3_to_row = {}
    for alias, iso3s in load_exiobase_region_members().items():
        a = alias.upper()
        if a in {"WA", "WE", "WF", "WL", "WM"}:  # short aliases only (skip the long names)
            for i in iso3s:
                iso3_to_row[i] = a
    iso3 = dict(zip(iso2s, _iso2_to_iso3(pd.Series(list(iso2s)))))
    return {c: (c if c in keep else iso3_to_row.get(iso3.get(c))) for c in iso2s}


def get_steel_supply_mix(
    api_url: str = DEFAULT_API,
    year: int = 2024,
    regions=None,
    source: str = "World Steel in Figures - crude steel by route",
) -> dict[str, dict[str, float]]:
    """Primary/secondary steel supply mix per EXIOBASE region, for `update_supply_mix`.

    Collapses the WSTEEL route production (BF-BOF + DRI-EAF = primary, scrap-EAF =
    secondary), maps ISO2 -> the table's regions (`regions` = `db.get_index('Region')`:
    the 44 EXIOBASE countries kept as-is, everyone else aggregated into their RoW group
    via MARIO's packaged membership), and returns
    ``{region: {STEEL_PRIMARY_ACT: p, STEEL_SECONDARY_ACT: s}}`` — the Activity-mix input:

        db.update_supply_mix(mix, level='Activity', commodities=[STEEL_COMMODITY],
                             scenario=..., rescale=True)

    Aggregation is on absolute Mt so RoW shares are production-weighted. Regions with no
    route data are absent (they keep their current table share).
    """
    frame = _get_csv(
        api_url, {"parameter": "Supply", "source": source, "period": str(year), "sort": "id"}
    )
    if frame.empty:
        raise ValueError(f"no Supply rows for source {source!r} in {year}")
    parts = frame["item_2"].map(split_item)  # [route_short, ISO2, period]
    df = pd.DataFrame(
        {
            "iso2": parts.str[1],
            "side": parts.str[0].map(
                {"BF.BOF": "primary", "DRI.EAF": "primary", "SCRAP.EAF": "secondary"}
            ),
            "mt": frame["value"].astype(float),
        }
    )
    df["region"] = df["iso2"].map(_iso2_to_exiobase_region(df["iso2"].unique(), regions))
    df = df.dropna(subset=["region"])
    agg = df.groupby(["region", "side"])["mt"].sum().unstack(fill_value=0.0)
    mix: dict[str, dict[str, float]] = {}
    for region, row in agg.iterrows():
        p, s = float(row.get("primary", 0.0)), float(row.get("secondary", 0.0))
        if p + s > 0:
            mix[region] = {
                STEEL_PRIMARY_ACT: round(p / (p + s), 6),
                STEEL_SECONDARY_ACT: round(s / (p + s), 6),
            }
    return mix


ALUMINIUM_COMMODITY = "Aluminium and aluminium products"

# EXIOBASE/nxsut commodity -> the HS6 prefixes (chapters/headings) that resolve to
# it, DERIVED from the nxbase HS22 -> CN26 -> NXS graph (the anchoring set by
# scripts/workbook/refine_trade_anchors.py). This is a materialisation of the
# intrinsic nxbase mapping for the commodities we update; regenerate the prefixes
# if the anchoring changes. Chemicals nec = chemical section VI minus fertilisers
# (ch 31); N-fertiliser = the nitrogen headings 3102.
_COMMODITY_HS: dict[str, tuple[str, ...]] = {
    STEEL_COMMODITY: ("72",),
    ALUMINIUM_COMMODITY: ("76",),
    "Chemicals nec": ("28", "29", "30", "32", "33", "34", "35", "36", "37"),
    "N-fertiliser": ("3102",),
    "Other non-metallic mineral products": ("68",),
    "Other non-ferrous metal products": ("81",),
}


_BACI_TRD_CACHE: dict = {}

# UN M49 codes BACI's own country_codes leaves without an ISO2 (a regional
# aggregate, or the "NA" it avoids writing for Namibia). Same fix as the nxbase
# m49_to_iso2 transform.
_M49_ISO2_FILE_OVERRIDES = {"490": "TW", "516": "NA"}


def _baci_frame(api_url, year, source, baci_path):
    """Normalised BACI frame (columns hs6, dest, origin, q), cached per run.

    ONE big pull, reused across every commodity in a run. PRIMARY source is the
    nxbase query API; if nxbase has no BACI (e.g. the public API, where BACI is
    visibility=local), it FALLS BACK to the raw CEPII files at ``baci_path`` — the
    download the user keeps locally, exactly like the EXIOBASE Hybrid flows and
    the EMBER release. So the pipeline stays reproducible from the public open API
    plus a local BACI file, without BACI ever being hosted. BACI has no public API
    (CEPII ships files); UN Comtrade has one but returns un-harmonised data.
    """
    key = (api_url, source, year, baci_path)
    if key in _BACI_TRD_CACHE:
        return _BACI_TRD_CACHE[key]
    frame = _baci_frame_from_api(api_url, year, source)
    if frame is None and baci_path:
        frame = _baci_frame_from_file(baci_path, year)
        if frame is not None:
            print(f"[nxbase] BACI not on the nxbase API; using local files at {baci_path}")
    if frame is None or frame.empty:
        raise ValueError(
            f"BACI unavailable for year<= {year}: not on nxbase ({api_url}) and no "
            "local fallback (set 'baci' in paths.yml to your CEPII BACI folder)"
        )
    _BACI_TRD_CACHE[key] = frame
    return frame


def _baci_frame_from_api(api_url, year, source):
    """Normalised BACI frame from the nxbase API, or None when nxbase has none."""
    if source is not None:
        raw = _get_csv(api_url, {"parameter": "Bilateral trade", "source": source})
    else:
        raw = None  # BACI lags the build year: use the latest vintage <= year
        for y in range(year, year - 3, -1):
            cand = f"BACI HS22 bilateral trade quantities {y}"
            f = _get_csv(api_url, {"parameter": "Bilateral trade", "source": cand})
            if not f.empty:
                raw = f
                if y != year:
                    print(f"[nxbase] BACI {year} unavailable, using {y}")
                break
    if raw is None or raw.empty:
        return None
    parts = raw["item_1"].map(split_item)  # c_<hs6>-<dest>-<period>
    return pd.DataFrame({
        "hs6": parts.str[0],
        "dest": parts.str[1],
        "origin": raw["item_2"].map(lambda s: split_item(s)[0]),  # s_<origin>
        "q": raw["value"].astype(float),
    })


def _baci_frame_from_file(baci_path, year):
    """Normalised BACI frame from the raw CEPII download (data + country_codes).

    Mirrors the nxbase baci_q recipe: M49->ISO2 from BACI's own country_codes
    (with the 490 "Other Asia, nes"->TW / 516 Namibia->NA overrides), HS6
    zero-padded, q>0. Uses the latest BACI_HS22_Y<yy> file <= ``year``. Unlike the
    nxbase subset this file carries ALL chapters, so every commodity resolves.
    """
    import glob
    import os

    cc = glob.glob(os.path.join(baci_path, "country_codes_*.csv"))
    if not cc:
        raise FileNotFoundError(f"no country_codes_*.csv in {baci_path}")
    codes = pd.read_csv(cc[0], dtype=str)
    m49_iso2 = {
        str(r["country_code"]).strip(): str(r.get("country_iso2") or "").strip()
        for _, r in codes.iterrows()
    }
    m49_iso2 = {k: v for k, v in m49_iso2.items() if v and v != "NA"}
    m49_iso2.update(_M49_ISO2_FILE_OVERRIDES)

    data_file = None
    for y in range(year, year - 3, -1):
        hits = glob.glob(os.path.join(baci_path, f"BACI_HS22_Y{y}_*.csv"))
        if hits:
            data_file = hits[0]
            if y != year:
                print(f"[nxbase] local BACI {year} unavailable, using {y}")
            break
    if not data_file:
        raise FileNotFoundError(f"no BACI_HS22_Y<= {year}_*.csv in {baci_path}")

    raw = pd.read_csv(data_file, usecols=["i", "j", "k", "q"])
    raw = raw[raw["q"] > 0]
    return pd.DataFrame({
        "hs6": raw["k"].map(lambda v: str(int(v)).zfill(6)),
        "dest": raw["j"].astype(str).map(m49_iso2.get),
        "origin": raw["i"].astype(str).map(m49_iso2.get),
        "q": raw["q"].astype(float),
    }).dropna(subset=["dest", "origin"])


def _baci_trade_mix(api_url, year, regions, hs_prefixes, source, baci_path):
    """Foreign sourcing shares per destination for the given HS6 prefixes.

    Reads the normalised BACI frame (nxbase API, or the local CEPII file
    fallback), keeps the HS6 codes under ``hs_prefixes`` (the chapters/headings
    that map to the target commodity in the nxbase graph), maps exporter/importer
    ISO2 -> the table's EXIOBASE regions (44 members kept, others RoW-aggregated),
    drops intra-region pairs, and returns ``{destination: {origin: share}}`` with
    each destination's FOREIGN origins summing to 1 (domestic diagonal omitted, so
    update_trade_mix preserves the base domestic share and rewrites only imports).
    """
    frame = _baci_frame(api_url, year, source, baci_path)
    df = frame[frame["hs6"].str.startswith(tuple(hs_prefixes))]
    if df.empty:
        raise ValueError(
            f"no BACI rows for HS prefixes {tuple(hs_prefixes)} "
            "(the nxbase subset / your local file has no such chapters)"
        )
    region = _iso2_to_exiobase_region(pd.unique(pd.concat([df["dest"], df["origin"]])), regions)
    df = df.assign(dest_r=df["dest"].map(region), origin_r=df["origin"].map(region))
    df = df.dropna(subset=["dest_r", "origin_r"])
    df = df[df["dest_r"] != df["origin_r"]]  # foreign only; domestic preserved by omission
    agg = df.groupby(["dest_r", "origin_r"])["q"].sum()
    mix: dict[str, dict[str, float]] = {}
    for dest_r, g in agg.groupby(level=0):
        total = g.sum()
        if total > 0:
            mix[dest_r] = {origin_r: round(v / total, 6) for (_, origin_r), v in g.items()}
    return mix


def get_trade_mix(
    api_url: str = DEFAULT_API,
    commodity: str | None = None,
    year: int = 2024,
    regions=None,
    source: str | None = None,
    baci_path: str | None = None,
) -> dict[str, dict[str, float]]:
    """Foreign sourcing shares per destination for one commodity, from BACI.

    Resolves the commodity's HS6 prefixes from the nxbase graph materialisation
    (`_COMMODITY_HS`) and returns ``{destination: {origin: share}}`` (foreign
    origins summing to 1, domestic omitted). Serves both paths:

    - **pooled** (steel/aluminium): ``items=pooled['supply']``,
      ``commodities=pooled['need']``;
    - **Isard** (chemicals, N-fertiliser, non-metallic/non-ferrous):
      ``items=[commodity]``, ``level='Commodity'``.

    Either way `update_trade_mix(..., rescale=True)` preserves the base domestic
    share and rewrites only the import sourcing. ``baci_path`` (the local CEPII
    BACI folder) is the fallback when the nxbase API has no BACI.
    """
    if commodity not in _COMMODITY_HS:
        raise KeyError(
            f"no HS mapping for {commodity!r}; add it to _COMMODITY_HS "
            "(derive the prefixes from the nxbase HS22->CN26->NXS graph)"
        )
    return _baci_trade_mix(api_url, year, regions, _COMMODITY_HS[commodity], source, baci_path)


def get_steel_trade_mix(api_url=DEFAULT_API, year=2024, regions=None, source=None, baci_path=None):
    """Foreign steel (HS ch.72) sourcing shares — thin wrapper over get_trade_mix."""
    return get_trade_mix(api_url, STEEL_COMMODITY, year, regions, source, baci_path)


def get_aluminium_trade_mix(api_url=DEFAULT_API, year=2024, regions=None, source=None, baci_path=None):
    """Foreign aluminium (HS ch.76) sourcing shares — thin wrapper over get_trade_mix."""
    return get_trade_mix(api_url, ALUMINIUM_COMMODITY, year, regions, source, baci_path)


def get_glass_recycled_share(
    api_url: str = DEFAULT_API,
    year: int = 2024,
    source: str = "FEVE / Close the Glass Loop - glass collection rate",
) -> dict[str, float]:
    """Per-country recycled/secondary glass share (0-1), `{ISO2: share}`.

    Queries `/data.csv?parameter=Market share&source=FEVE&period=<year>`; the value is
    the FEVE collection-for-recycling rate (a proxy for the cullet supply share). EU-only.
    Scale by the EU-average cullet content (~0.535) at the call site if you want the
    absolute secondary share; use it directly for a relative country modulation.
    """
    frame = _get_csv(
        api_url,
        {"parameter": "Market share", "source": source, "period": str(year), "sort": "id"},
    )
    if frame.empty:
        raise ValueError(f"no Market share rows for source {source!r} in {year}")
    region = frame["item_1"].map(lambda s: split_item(s)[1])  # c_GLAW-<ISO2>-<period>
    return dict(zip(region, frame["value"].astype(float)))


# base-table (EXIOBASE) activities competing to supply the glass commodity
GLASS_PRIMARY_ACT = "Manufacture of glass and glass products"
GLASS_SECONDARY_ACT = "Re-processing of secondary glass into new glass"
GLASS_COMMODITY = "Glass and glass products"
GLASS_CULLET_EU = 0.5355  # FEVE EU-average recycled (cullet) content of container glass


def get_glass_supply_mix(
    api_url: str = DEFAULT_API,
    year: int = 2024,
    regions=None,
    cullet_eu: float = GLASS_CULLET_EU,
    source: str = "FEVE / Close the Glass Loop - glass collection rate",
) -> dict[str, dict[str, float]]:
    """Primary/secondary glass supply mix per EXIOBASE region (EU-only), for `update_supply_mix`.

    FEVE gives the per-country **collection rate** (waste side); the **cullet content** (supply
    side, what the SUT split needs) is only published as an EU average (`cullet_eu` ~0.535). So
    the per-country secondary (cullet) share is approximated by anchoring to that EU average and
    modulating by the country's collection performance relative to the EU mean:

        secondary_c = min(cullet_eu * collection_c / mean(collection), 0.95)

    Returns ``{region: {GLASS_PRIMARY_ACT: 1-s, GLASS_SECONDARY_ACT: s}}`` for
    ``db.update_supply_mix(mix, level='Activity', commodities=[GLASS_COMMODITY], rescale=True)``.
    EU-only: non-FEVE regions keep their base share; the bottle-reuse activity is left untouched.
    """
    rates = get_glass_recycled_share(api_url, year, source)  # {ISO2: collection_rate}
    if not rates:
        raise ValueError(f"no glass rows for source {source!r} in {year}")
    eu_avg = sum(rates.values()) / len(rates)
    region_of = _iso2_to_exiobase_region(list(rates), regions)
    by_region: dict[str, list[float]] = {}
    for iso2, rate in rates.items():
        region = region_of.get(iso2)
        if region is None:
            continue
        by_region.setdefault(region, []).append(min(cullet_eu * rate / eu_avg, 0.95))
    return {
        region: {
            GLASS_PRIMARY_ACT: round(1 - (s := sum(v) / len(v)), 6),
            GLASS_SECONDARY_ACT: round(s, 6),
        }
        for region, v in by_region.items()
    }


def get_plastics_recycled_share(
    api_url: str = DEFAULT_API,
    year: int = 2019,
    source: str = "OECD Global Plastics Outlook - world secondary share",
) -> float:
    """World secondary (recycled) plastics use share (0-1) for `year` (site LXX).

    The OECD split is world-only; apply this uniformly across regions (the per-region
    OECD-macro-region rate is a documented follow-up). OECD baseline stops at 2019.
    """
    frame = _get_csv(
        api_url,
        {"parameter": "Market share", "source": source, "period": str(year), "sort": "id"},
    )
    if frame.empty:
        raise ValueError(f"no Market share rows for source {source!r} in {year}")
    return float(frame["value"].iloc[0])


# Pipeline-side cluster: nxbase-resolved concept -> the base table's label.
_ADD_SECTORS_REVERSE_CLUSTER = {
    "Carbon dioxide, fossil": "CO2",
    "Methane, fossil": "CH4",
    "Nitrous oxide": "N2O",
    "Operating surplus: Consumption of fixed capital": "CAPEX",
    "Coke Oven Coke": "Coke",
}
# Base-table remap applied to the master's DB Item column. add_sectors runs
# AFTER aggregate_ee, where the electricity generation commodities are folded
# into a single grid "Electricity" — so the ETE master's consolidated
# "Electricity" / "Electricity RES" both point at "Electricity". (A finer
# electricity attachment is a known later refinement.)
_ADD_SECTORS_DBITEM_REMAP = {
    "Electricity": "Electricity",
    "Electricity RES": "Electricity",
}


def build_add_sectors_master(
    template_path: str,
    out_path: str,
    api_url: str = DEFAULT_API,
    source: str = "Ghezzi et al. 2026 - steel & H2 inventory",
) -> str:
    """Write an add_sectors master driven by the nxbase recipe.

    The **recipe** (per-route quantities) comes from nxbase; the **base-DB
    attachment** — the GLOBAL/market-share placement and the Regions cluster
    sheet — is kept from ``template_path`` (the original master). Two ETE-legacy
    fixes are applied for nxsut's post-``aggregate_ee`` base: the DB Item column
    remaps electricity to the single grid "Electricity", and the ETE
    "Commodities Clusters" sheet (which groups generation-split electricity
    commodities that do not exist here — and collide with the aggregated EMBER
    activity labels) is cleared. add_sectors must run AFTER aggregate_ee.
    Returns ``out_path``.
    """
    import openpyxl

    recipe = get_add_sectors_recipe(api_url, source)
    # (activity, base-table label) -> quantity from nxbase
    qty: dict[tuple[str, str], float] = {}
    for _, r in recipe[recipe["item_type"] != "output"].iterrows():
        label = _ADD_SECTORS_REVERSE_CLUSTER.get(r["input"], r["input"])
        qty[(r["route"].strip(), label.strip())] = r["quantity"]

    wb = openpyxl.load_workbook(template_path)
    # Drop the ETE electricity Commodities Clusters (Coal/Hydro/Solar... are not
    # commodities in the aggregated base, and 'Coal' collides with the EMBER
    # activity label -> copy_from_parent KeyError). Keep the header row only.
    cc = wb["Commodities Clusters"]
    if cc.max_row > 1:
        cc.delete_rows(2, cc.max_row - 1)
    ms = wb["Master"]
    mh = [c.value for c in next(ms.iter_rows(min_row=1, max_row=1))]
    mi = {h: i for i, h in enumerate(mh)}
    sheet_to_act = {
        ms.cell(row=rr, column=mi["Inventory sheet"] + 1).value: ms.cell(row=rr, column=mi["Activity"] + 1).value
        for rr in range(2, ms.max_row + 1)
        if ms.cell(row=rr, column=mi["Inventory sheet"] + 1).value
    }
    struct = {"Master", "Commodities Clusters", "Regions Clusters", "DB units"}
    for sheet in wb.sheetnames:
        if sheet in struct:
            continue
        ws = wb[sheet]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {h: i for i, h in enumerate(hdr)}
        act = str(sheet_to_act.get(sheet, "")).strip()
        for rr in range(2, ws.max_row + 1):
            label = ws.cell(row=rr, column=idx["Input"] + 1).value
            if label is None:
                continue
            # remap the DB Item for nxsut's aggregated base
            dbi_cell = ws.cell(row=rr, column=idx["DB Item"] + 1)
            if dbi_cell.value in _ADD_SECTORS_DBITEM_REMAP:
                dbi_cell.value = _ADD_SECTORS_DBITEM_REMAP[dbi_cell.value]
            # overwrite quantity from nxbase where the recipe carries this input
            key = (act, str(label).strip())
            if key in qty:
                ws.cell(row=rr, column=idx["Quantity"] + 1).value = qty[key]
    wb.save(out_path)
    return out_path

